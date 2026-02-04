import sys
import sqlite3
import shutil
import atexit
from pathlib import Path
from datetime import datetime
from PySide6.QtWidgets import (
    QApplication, QWidget, QMainWindow, QListWidget, QStackedWidget,
    QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLineEdit, QTextEdit,
    QFormLayout, QTableWidget, QTableWidgetItem, QMessageBox, QInputDialog,
    QDialog, QListWidgetItem, QFileDialog, QCheckBox, QComboBox, QSpinBox,
    QHeaderView, QDoubleSpinBox
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont
from ui.dashboard import Dashboard

DB_PATH = Path("selection.db")
ATTACH_DIR = Path("attachments")
ATTACH_DIR.mkdir(exist_ok=True)

class ContributionDialog(QDialog):
    def __init__(self, evaluation_id: int, team_id: int, parent=None):
        super().__init__(parent)
        self.evaluation_id = evaluation_id
        self.team_id = team_id
        self.setWindowTitle(f"Contribuição Individual - Avaliação {evaluation_id}")
        self.resize(600, 400)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Ajuste o peso de contribuição (0.1 a 1.2). Padrão: 1.0"))

        self.members_table = QTableWidget(0, 4)
        self.members_table.setHorizontalHeaderLabels(["ID", "Nome", "Peso", "Obs. Interna"])
        self.members_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.members_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        layout.addWidget(self.members_table)

        save_btn = QPushButton("Salvar Contribuições")
        save_btn.setObjectName("primary")
        save_btn.clicked.connect(self.save_contributions)
        layout.addWidget(save_btn)

        self.load_members()

    def load_members(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("""
            SELECT c.id, c.name FROM candidates c
            JOIN team_members tm ON c.id = tm.candidate_id
            WHERE tm.team_id = ? ORDER BY c.name
        """, (self.team_id,))
        members = c.fetchall()
        conn.close()
        self.members_table.setRowCount(len(members))
        for r, (member_id, name) in enumerate(members):
            self.members_table.setItem(r, 0, QTableWidgetItem(str(member_id)))
            self.members_table.setItem(r, 1, QTableWidgetItem(name))

            weight_spinbox = QDoubleSpinBox()
            weight_spinbox.setRange(0.1, 1.2)
            weight_spinbox.setSingleStep(0.1)
            weight_spinbox.setValue(1.0)
            self.members_table.setCellWidget(r, 2, weight_spinbox)

            note_edit = QLineEdit()
            self.members_table.setCellWidget(r, 3, note_edit)
        
        # Make ID column read-only
        for r in range(self.members_table.rowCount()):
            item = self.members_table.item(r, 0)
            if item:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item = self.members_table.item(r, 1)
            if item:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)


    def save_contributions(self):
        contributions = []
        for r in range(self.members_table.rowCount()):
            member_id = int(self.members_table.item(r, 0).text())
            weight = self.members_table.cellWidget(r, 2).value()
            note = self.members_table.cellWidget(r, 3).text().strip()
            contributions.append((self.evaluation_id, member_id, weight, note))

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        try:
            c.executemany("""
                INSERT INTO member_contribution (evaluation_id, member_id, weight, note)
                VALUES (?, ?, ?, ?)
            """, contributions)
            conn.commit()
            QMessageBox.information(self, "Sucesso", "Contribuições individuais salvas.")
            audit('member_contribution_save', f'evaluation_id={self.evaluation_id}, count={len(contributions)}')
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Erro de Banco de Dados", f"Não foi possível salvar as contribuições: {e}")
        finally:
            conn.close()


# -------------------------------
# MIGRAÇÃO DE BANCO (schema v1+)
# -------------------------------
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    # user_version indica versão do schema
    cur.execute("PRAGMA user_version")
    ver = cur.fetchone()[0] or 0

    # v0 -> v1 (tabelas básicas)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS candidates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT,
            notes TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS teams (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS team_members (
            team_id INTEGER,
            candidate_id INTEGER,
            PRIMARY KEY(team_id, candidate_id)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS evaluations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            team_id INTEGER,
            judge TEXT,
            immersion INTEGER,
            development INTEGER,
            presentation INTEGER,
            notes TEXT,
            hidden_score REAL DEFAULT 0
        )
    """)
    cur.execute("PRAGMA user_version = 1")

    # v1 -> v2: competição na equipe + veteranos + sessões + presença
    if ver < 2:
        try:
            cur.execute("ALTER TABLE teams ADD COLUMN competition TEXT DEFAULT 'OBR'")
        except sqlite3.OperationalError:
            pass
        try:
            cur.execute("ALTER TABLE teams ADD COLUMN is_veteran INTEGER DEFAULT 0")
        except sqlite3.OperationalError:
            pass
        cur.execute("""
            CREATE TABLE IF NOT EXISTS training_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                training_session_id INTEGER,
                team_id INTEGER,
                present INTEGER DEFAULT 1,
                notes TEXT
            )
        """)
        cur.execute("PRAGMA user_version = 2")

    # v2 -> v3: diário e anexos
    if ver < 3:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS diary_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                team_id INTEGER,
                title TEXT,
                content TEXT,
                created_at TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                diary_entry_id INTEGER,
                file_path TEXT,
                original_name TEXT,
                mime_type TEXT
            )
        """)
        cur.execute("PRAGMA user_version = 3")

    # v3 -> v4: sessão e comentário na avaliação
    if ver < 4:
        try:
            cur.execute("ALTER TABLE evaluations ADD COLUMN training_session_id INTEGER")
        except sqlite3.OperationalError:
            pass
        try:
            cur.execute("ALTER TABLE evaluations ADD COLUMN comment TEXT")
        except sqlite3.OperationalError:
            pass
        cur.execute("PRAGMA user_version = 4")

    # v4 -> v5: pesos internos configuráveis
    if ver < 5:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS internal_weights (
                name TEXT PRIMARY KEY,
                weight REAL NOT NULL
            )
        """)
        for name, w in [('immersion', 0.3), ('development', 0.5), ('presentation', 0.2)]:
            cur.execute("INSERT OR IGNORE INTO internal_weights(name,weight) VALUES(?,?)", (name, w))
        cur.execute("PRAGMA user_version = 5")

    # v5 -> v6: settings (PIN hash)
    if ver < 6:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """)
        try:
            import hashlib
            h = hashlib.sha256(b"1234").hexdigest()
            cur.execute("INSERT OR IGNORE INTO settings(key,value) VALUES('admin_hash',?)", (h,))
            cur.execute("INSERT OR IGNORE INTO settings(key,value) VALUES('process_status',?)", ('ABERTO',))
        except Exception:
            pass
        cur.execute("PRAGMA user_version = 6")

    # v6 -> v7: cpf, phone, grade
    if ver < 7:
        try:
            cur.execute("ALTER TABLE candidates ADD COLUMN cpf TEXT")
        except sqlite3.OperationalError:
            pass
        try:
            cur.execute("ALTER TABLE candidates ADD COLUMN phone TEXT")
        except sqlite3.OperationalError:
            pass
        try:
            cur.execute("ALTER TABLE candidates ADD COLUMN grade TEXT")
        except sqlite3.OperationalError:
            pass
        cur.execute("PRAGMA user_version = 7")

    # v7 -> v8: member_contribution table
    if ver < 8:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS member_contribution (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                evaluation_id INTEGER,
                member_id INTEGER,
                weight REAL,
                note TEXT,
                FOREIGN KEY(evaluation_id) REFERENCES evaluations(id),
                FOREIGN KEY(member_id) REFERENCES candidates(id)
            )
        """)
        cur.execute("PRAGMA user_version = 8")

    # v8 -> v9: soft delete para avaliações
    if ver < 9:
        try:
            cur.execute("ALTER TABLE evaluations ADD COLUMN is_active INTEGER DEFAULT 1")
        except sqlite3.OperationalError:
            pass
        try:
            cur.execute("ALTER TABLE evaluations ADD COLUMN deleted_at TEXT")
        except sqlite3.OperationalError:
            pass
        try:
            cur.execute("ALTER TABLE evaluations ADD COLUMN delete_reason TEXT")
        except sqlite3.OperationalError:
            pass
        cur.execute("PRAGMA user_version = 9")

    conn.commit()
    conn.close()

# --------------------------------
# ESTILOS E UTILITÁRIOS
# --------------------------------
def dark_stylesheet() -> str:
    return """
    QWidget {
        background-color: #121212;
        color: #e0e0e0;
        font-family: "Segoe UI", Roboto, Arial, sans-serif;
    }
    QMainWindow, QFrame, QStackedWidget { background-color: #121212; }
    QListWidget { background-color: #141414; border: 1px solid #2a2a2a; }
    QListWidget::item { padding: 8px 10px; }
    QListWidget::item:selected { background: #1e88e5; color: #ffffff; }
    QPushButton {
        background-color: #1f1f1f;
        border: 1px solid #2b2b2b;
        color: #e0e0e0;
        padding: 6px 10px;
        border-radius: 6px;
    }
    QPushButton:hover { background-color: #2a2a2a; }
    QPushButton:pressed { background-color: #164f77; }
    QLineEdit, QTextEdit, QPlainTextEdit, QTableWidget {
        background-color: #0f0f0f;
        border: 1px solid #2b2b2b;
        color: #e0e0e0;
    }
    QHeaderView::section { background-color: #1a1a1a; color: #e0e0e0; padding: 4px; border: 0px; }
    QTableWidget::item:selected { background: #263238; color: #ffffff; }
    QLabel { color: #e0e0e0; }
    QComboBox { background-color: #0f0f0f; border: 1px solid #2b2b2b; color: #e0e0e0; }
    """

def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Settings helpers and audit
def get_setting(key, default=None):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    try:
        cur.execute("SELECT value FROM settings WHERE key=?", (key,))
        r = cur.fetchone()
        return r[0] if r else default
    finally:
        conn.close()

def set_setting(key, value):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    try:
        cur.execute("REPLACE INTO settings (key,value) VALUES (?,?)", (key, str(value)))
        conn.commit()
    finally:
        conn.close()

def get_process_status():
    """Retorna o status atual do processo seletivo (ABERTO/ENCERRADO)."""
    return get_setting('process_status', 'ABERTO')

def set_process_status(status: str):
    """Define o status do processo seletivo (ABERTO/ENCERRADO)."""
    set_setting('process_status', status)
    audit('process_status_change', f'status={status}')

def audit(action, details=""):
    ts = datetime.now().isoformat(sep=' ', timespec='seconds')
    with open('audit.log', 'a', encoding='utf-8') as f:
        f.write(f"[{ts}] {action} {details}\n")

# --- HELPERS PARA COMBOBOXES (IDs -> labels) ---
def fetch_teams():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("SELECT id, name FROM teams ORDER BY name ASC")
    rows = c.fetchall()
    conn.close()
    return rows

def fetch_sessions():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("SELECT id, date, start_time, end_time FROM training_sessions ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows

def fill_team_combobox(cb: QComboBox):
    cb.clear()
    for tid, tname in fetch_teams():
        cb.addItem(f"{tid} - {tname}", tid)

def fill_session_combobox(cb: QComboBox):
    cb.clear()
    for sid, d, s, e in fetch_sessions():
        cb.addItem(f"{sid} - {d} ({s}-{e})", sid)

# Backup automático
def backup_snapshot(prefix="auto"):
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    dst = Path(f"{prefix}_backup_selection_{ts}.db")
    try:
        shutil.copy2(DB_PATH, dst)
        with open('backup.log', 'a', encoding='utf-8') as f:
            f.write(f"[{now_str()}] backup -> {dst}\n")
    except Exception as e:
        with open('backup.log', 'a', encoding='utf-8') as f:
            f.write(f"[{now_str()}] backup failed: {e}\n")

# -------------------------------
# JANELA PRINCIPAL
# -------------------------------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Processo Seletivo — RobotO1e")
        self.resize(1200, 800)
        main = QWidget()
        self.setCentralWidget(main)
        hb = QHBoxLayout(main)

        # Status do Processo Seletivo
        self.status = self.statusBar()

        self.process_status_label = QLabel("Processo Seletivo: ABERTO")
        self.process_status_label.setStyleSheet("""
            QLabel {
                color: #22c55e;
                font-weight: bold;
                padding-left: 8px;
            }
        """)

        self.status.addPermanentWidget(self.process_status_label)
        
        self.sidebar = QListWidget()
        self.sidebar.addItems([
            "Inscrições", "Equipes", "Sessões", "Presença",
            "Avaliações", "Diário de Bordo", "Sobre", "Dashboard", "Admin (oculto)"
        ])
        self.sidebar.currentRowChanged.connect(self.on_nav)
        hb.addWidget(self.sidebar, 1)

        self.stack = QStackedWidget()
        hb.addWidget(self.stack, 4)

        # Páginas
        self.stack.addWidget(self.page_registrations())   # 0
        self.stack.addWidget(self.page_teams())           # 1
        self.stack.addWidget(self.page_sessions())        # 2
        self.stack.addWidget(self.page_attendance())      # 3
        self.stack.addWidget(self.page_evaluations())     # 4
        self.stack.addWidget(self.page_diary())           # 5
        self.stack.addWidget(self.page_about())           # 6
        self.stack.addWidget(self.page_dashboard())       # 7
        self.stack.addWidget(self.page_admin())           # 8

        self.sidebar.setCurrentRow(0)

    def _get_selected_id(self, table: QTableWidget, label: str, col: int = 0):
        row = table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Ação", f"Selecione um registro de {label} primeiro.")
            return None
        item = table.item(row, col)
        if not item:
            QMessageBox.warning(self, "Ação", f"Não foi possível ler o ID de {label}.")
            return None
        try:
            return int(item.text())
        except ValueError:
            QMessageBox.warning(self, "Ação", f"ID inválido para {label}.")
            return None

    def on_nav(self, idx):
        # Protege Admin por PIN (hash na tabela settings)
        if idx == 8:
            pin, ok = QInputDialog.getText(self, "PIN de Acesso", "Insira o PIN da banca:", QLineEdit.Password)
            if not ok:
                self.sidebar.setCurrentRow(0)
                return
            import hashlib
            h = hashlib.sha256(pin.encode()).hexdigest()
            stored = get_setting('admin_hash', '')
            if h != stored:
                QMessageBox.warning(self, "Acesso negado", "PIN incorreto.")
                self.sidebar.setCurrentRow(0)
                return
        self.stack.setCurrentIndex(idx)
        closed = get_process_status() == "ENCERRADO"
        allowed = {6, 7, 8}  # Sobre, Dashboard, Admin
        if closed and idx not in allowed:
            QMessageBox.information(self, "Processo Encerrado",
                                    "As abas de cadastro e edição estão bloqueadas. Use Sobre, Dashboard ou Admin.")
            self.sidebar.setCurrentRow(6)  # Redireciona para Sobre
            self.stack.setCurrentIndex(6)
            return
    
    
    def apply_process_lockdown(self):
        closed = get_process_status() == "ENCERRADO"
        allowed = {6, 7, 8}
        for i in range(self.sidebar.count()):
            item = self.sidebar.item(i)
            if closed and i not in allowed:
                # Desabilita se fechado
                item.setFlags(item.flags() & ~Qt.ItemIsEnabled)
            else:
                # Reabilita se aberto
                item.setFlags(item.flags() | Qt.ItemIsEnabled)

    def update_process_status_display(self):
        status = get_process_status()
        self.process_status_label.setText(f"Processo Seletivo: <b>{status}</b>")
        if status == "ENCERRADO":
            self.process_status_label.setStyleSheet("color: red;")
            self.apply_process_lockdown()
        else:
            self.process_status_label.setStyleSheet("color: green;")

    # --------------------------
    # PÁGINA: INSCRIÇÕES
    # --------------------------
    def page_registrations(self):
        w = QWidget()
        layout = QVBoxLayout(w)

        form = QFormLayout()
        self.name_in = QLineEdit()
        self.email_in = QLineEdit()
        self.cpf_in = QLineEdit()
        self.phone_in = QLineEdit()
        self.grade_in = QLineEdit()
        self.notes_in = QTextEdit()
        form.addRow("Nome:", self.name_in)
        form.addRow("E-mail:", self.email_in)
        form.addRow("CPF:", self.cpf_in)
        form.addRow("Telefone:", self.phone_in)
        form.addRow("Série:", self.grade_in)
        form.addRow("Observações:", self.notes_in)
        add_btn = QPushButton("Adicionar candidato")
        add_btn.setObjectName("primary")
        add_btn.clicked.connect(self.add_candidate)
        layout.addLayout(form)
        layout.addWidget(add_btn)

        self.cand_table = QTableWidget(0, 6)
        self.cand_table.setHorizontalHeaderLabels(["ID", "Nome", "E-mail", "CPF", "Telefone", "Série"])
        self.cand_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.cand_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.cand_table.setSortingEnabled(True)
        self.cand_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        layout.addWidget(self.cand_table)

        # busca
        search_row = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Buscar por nome ou e-mail")
        self.search_input.textChanged.connect(self.load_candidates)
        search_row.addWidget(QLabel("Buscar:"))
        search_row.addWidget(self.search_input)
        layout.addLayout(search_row)

        btns = QHBoxLayout()
        refresh_c = QPushButton("Atualizar")
        refresh_c.setObjectName("primary")
        refresh_c.clicked.connect(self.load_candidates)
        view_c = QPushButton("Ver/Editar")
        view_c.setObjectName("danger")
        view_c.clicked.connect(self.view_selected_candidate)
        delete_c = QPushButton("Remover")
        delete_c.setObjectName("danger")
        delete_c.clicked.connect(self.delete_selected_candidate)
        import_c = QPushButton("Importar CSV/XLSX")
        import_c.setObjectName("primary")
        import_c.clicked.connect(self.import_candidates_csv)
        btns.addWidget(refresh_c)
        btns.addWidget(view_c)
        btns.addWidget(delete_c)
        btns.addWidget(import_c)
        layout.addLayout(btns)

        self.load_candidates()
        return w

    def add_candidate(self):
        name = self.name_in.text().strip()
        email = self.email_in.text().strip()
        cpf = self.cpf_in.text().strip()
        phone = self.phone_in.text().strip()
        grade = self.grade_in.text().strip()
        notes = self.notes_in.toPlainText().strip()
        if not name:
            QMessageBox.warning(self, "Erro", "Nome é obrigatório")
            return
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("INSERT INTO candidates (name,email,notes,cpf,phone,grade) VALUES (?,?,?,?,?,?)", (name, email, notes, cpf, phone, grade))
        conn.commit()
        conn.close()
        self.name_in.clear(); self.email_in.clear(); self.notes_in.clear()
        self.cpf_in.clear(); self.phone_in.clear(); self.grade_in.clear()
        self.load_candidates()

    def load_candidates(self):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        q = "SELECT id,name,email,cpf,phone,grade FROM candidates"
        params = ()
        term = getattr(self, 'search_input', None)
        if term and term.text().strip():
            s = '%' + term.text().strip() + '%'
            q += " WHERE name LIKE ? OR email LIKE ? OR cpf LIKE ? OR phone LIKE ? OR grade LIKE ?"
            params = (s, s, s, s, s)
        q += " ORDER BY id DESC"
        cur.execute(q, params)
        rows = cur.fetchall()
        conn.close()
        self.cand_table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                self.cand_table.setItem(r, c, QTableWidgetItem(str(val)))

    def delete_selected_candidate(self):
        sel = self.cand_table.currentRow()
        if sel < 0:
            QMessageBox.warning(self, "Erro", "Selecione um candidato para remover")
            return
        item = self.cand_table.item(sel,0)
        if not item:
            return
        cid = int(item.text())
        ok = QMessageBox.question(self, "Confirmar", f"Remover candidato {cid}? Esta ação é irreversível.")
        if ok != QMessageBox.Yes:
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("DELETE FROM team_members WHERE candidate_id=?", (cid,))
        c.execute("DELETE FROM candidates WHERE id=?", (cid,))
        conn.commit()
        conn.close()
        self.load_candidates()

    def view_selected_candidate(self):
        sel = self.cand_table.currentRow()
        if sel < 0:
            QMessageBox.warning(self, "Erro", "Selecione um candidato")
            return
        item = self.cand_table.item(sel,0)
        if not item:
            return
        cid = int(item.text())
        dlg = CandidateDialog(cid, parent=self)
        dlg.exec()
        self.load_candidates()

    def import_candidates_csv(self):
        fn, _ = QFileDialog.getOpenFileName(self, "Importar arquivo (CSV/XLSX)", "", "Planilhas (*.csv *.xlsx);;CSV (*.csv);;Excel (*.xlsx)")
        if not fn:
            return
        p = Path(fn)
        inserted = 0
        skipped = 0
        # XLSX
        if p.suffix.lower() in ('.xlsx', '.xls'):
            try:
                from openpyxl import load_workbook
            except Exception as e:
                QMessageBox.warning(self, 'Erro', f'openpyxl não disponível: {e}')
                return
            wb = load_workbook(fn, read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            headers = []
            data_rows = []
            if rows:
                first = rows[0]
                if any(isinstance(x, str) for x in first):
                    headers = [str(x).strip() if x is not None else '' for x in first]
                    data_rows = rows[1:]
                else:
                    data_rows = rows
            preview = [tuple(r) for r in data_rows[:10]]
            dlg = ImportPreviewDialog(headers, preview, parent=self)
            if dlg.exec() != QDialog.Accepted:
                QMessageBox.information(self, 'Importar Excel', 'Importação cancelada')
                return
            idx_name, idx_email, idx_notes, skip_dup, idx_cpf, idx_phone, idx_grade = dlg.mapping_indices()
            conn = sqlite3.connect(DB_PATH); c = conn.cursor()
            for r in data_rows:
                if all(val is None or str(val).strip() == '' for val in r):
                    skipped += 1
                    continue
                def get_idx(r, idx):
                    try:
                        v = r[idx]
                        return str(v).strip() if v is not None else ''
                    except Exception:
                        return ''
                name = get_idx(r, idx_name)
                email = get_idx(r, idx_email)
                notes = get_idx(r, idx_notes)
                cpf = get_idx(r, idx_cpf)
                phone = get_idx(r, idx_phone)
                grade = get_idx(r, idx_grade)
                if not name.strip():
                    skipped += 1; continue
                if skip_dup:
                    c.execute("SELECT id FROM candidates WHERE name=? AND (email=? OR ?='')", (name, email, email))
                    if c.fetchone():
                        skipped += 1; continue
                try:
                    c.execute("INSERT INTO candidates (name,email,notes,cpf,phone,grade) VALUES (?,?,?,?,?,?)", (name, email, notes, cpf, phone, grade))
                    inserted += 1
                except Exception:
                    skipped += 1
            conn.commit(); conn.close()
            self.load_candidates()
            QMessageBox.information(self, 'Importar Excel', f'Concluída: {inserted} inseridos, {skipped} ignorados')
            audit('import_xlsx', f'file={p.name}, inserted={inserted}, skipped={skipped}')
            return
        # CSV
        import csv
        try:
            f = open(fn, newline='', encoding='utf-8')
        except UnicodeDecodeError:
            f = open(fn, newline='', encoding='latin-1')
        with f:
            reader = csv.DictReader(f)
            rows = []
            if reader.fieldnames:
                for r in reader:
                    rows.append(r)
                headers = list(rows[0].keys()) if rows else []
                preview_rows = [tuple(r.get(h, '') for h in headers) for r in rows[:10]]
            else:
                f.seek(0)
                f2 = csv.reader(f)
                rows = []
                for r in f2:
                    if not r: continue
                    rows.append({'__raw__': r})
                headers = []
                preview_rows = [tuple(r['__raw__']) for r in rows[:10]]
            dlg = ImportPreviewDialog(headers, preview_rows, parent=self)
            if dlg.exec() != QDialog.Accepted:
                QMessageBox.information(self, 'Importar CSV', 'Importação cancelada')
                return
            idx_name, idx_email, idx_notes, skip_dup, idx_cpf, idx_phone, idx_grade = dlg.mapping_indices()
            conn = sqlite3.connect(DB_PATH); c = conn.cursor()
            for r in rows:
                if '__raw__' in r:
                    raw = r['__raw__']
                    if all(val.strip() == '' for val in raw):
                        skipped += 1
                        continue
                    name = raw[idx_name].strip() if len(raw) > idx_name else ''
                    email = raw[idx_email].strip() if len(raw) > idx_email else ''
                    notes = raw[idx_notes].strip() if len(raw) > idx_notes else ''
                    cpf = raw[idx_cpf].strip() if len(raw) > idx_cpf else ''
                    phone = raw[idx_phone].strip() if len(raw) > idx_phone else ''
                    grade = raw[idx_grade].strip() if len(raw) > idx_grade else ''
                else:
                    if all(val.strip() == '' for val in r.values()):
                        skipped += 1
                        continue
                    def get_by_index(d, idx):
                        try:
                            key = list(d.keys())[idx]
                            return str(d.get(key,'')).strip()
                        except Exception:
                            return ''
                    name = get_by_index(r, idx_name)
                    email = get_by_index(r, idx_email)
                    notes = get_by_index(r, idx_notes)
                    cpf = get_by_index(r, idx_cpf)
                    phone = get_by_index(r, idx_phone)
                    grade = get_by_index(r, idx_grade)
                if not name.strip():
                    skipped += 1; continue
                if skip_dup:
                    c.execute("SELECT id FROM candidates WHERE name=? AND (email=? OR ?='')", (name, email, email))
                    if c.fetchone():
                        skipped += 1; continue
                try:
                    c.execute("INSERT INTO candidates (name,email,notes,cpf,phone,grade) VALUES (?,?,?,?,?,?)", (name, email, notes, cpf, phone, grade))
                    inserted += 1
                except Exception:
                    skipped += 1
            conn.commit(); conn.close()
            self.load_candidates()
            QMessageBox.information(self, 'Importar CSV', f'Concluída: {inserted} inseridos, {skipped} ignorados')
            audit('import_csv', f'file={Path(fn).name}, inserted={inserted}, skipped={skipped}')

    # AUTO-ATRIBUIÇÃO (round-robin por tamanho)
    def auto_assign_dialog(self):
        size, ok = QInputDialog.getInt(self, 'Auto-atribuir', 'Tamanho por equipe (ex: 3):', 3, 1, 100)
        if not ok:
            return
        self.auto_assign_by_size(size)

    def auto_assign_by_size(self, size:int):
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute("SELECT id FROM candidates WHERE id NOT IN (SELECT candidate_id FROM team_members)")
        unassigned = [r[0] for r in c.fetchall()]
        if not unassigned:
            QMessageBox.information(self, 'Auto-atribuir', 'Nenhum candidato sem equipe')
            conn.close(); return
        import math
        needed = math.ceil(len(unassigned) / size)
        c.execute("SELECT COUNT(*) FROM teams")
        existing = c.fetchone()[0]
        created = []
        for i in range(max(0, needed - existing)):
            name = f"AutoTeam_{now_str()}_{i}"
            c.execute("INSERT INTO teams(name,competition,is_veteran) VALUES(?,?,?)", (name, 'OBR', 0))
            created.append(c.lastrowid)
        c.execute("SELECT id FROM teams ORDER BY id ASC LIMIT ?", (needed,))
        team_ids = [r[0] for r in c.fetchall()]
        for idx, cid in enumerate(unassigned):
            tid = team_ids[idx % len(team_ids)]
            c.execute("INSERT OR IGNORE INTO team_members(team_id,candidate_id) VALUES(?,?)", (tid, cid))
        conn.commit(); conn.close()
        self.load_teams(); self.load_candidates()
        QMessageBox.information(self, 'Auto-atribuir', f'Atribuídos {len(unassigned)} candidatos em {len(team_ids)} equipes')
        audit('auto_assign', f'size={size},assigned={len(unassigned)}')

    # --------------------------
    # PÁGINA: EQUIPES
    # --------------------------
    def page_teams(self):
        w = QWidget()
        v = QVBoxLayout(w)
        form = QFormLayout()
        self.team_name_in = QLineEdit()
        self.team_comp_in = QComboBox(); self.team_comp_in.addItems(["OBR", "TBR", "CCBB"])
        self.team_vet_in = QCheckBox("Veteranos?")
        form.addRow("Nome da equipe:", self.team_name_in)
        form.addRow("Competição:", self.team_comp_in)
        form.addRow(self.team_vet_in)
        add_team = QPushButton("Criar equipe")
        add_team.clicked.connect(self.create_team)
        v.addLayout(form)
        v.addWidget(add_team)
        self.team_table = QTableWidget(0, 4)
        self.team_table.setHorizontalHeaderLabels(["ID", "Nome", "Competição", "Veteranos"])
        self.team_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.team_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.team_table.setSortingEnabled(True)
        self.team_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        v.addWidget(self.team_table)
        btn_row = QHBoxLayout()
        refresh = QPushButton("Atualizar")
        refresh.clicked.connect(self.load_teams)
        edit_team = QPushButton("Editar equipe")
        edit_team.clicked.connect(self.edit_selected_team)
        manage = QPushButton("Gerenciar membros")
        manage.clicked.connect(self.open_manage_dialog)
        delete_team = QPushButton("Remover equipe")
        delete_team.clicked.connect(self.delete_selected_team)
        auto_btn = QPushButton("Auto-atribuir")
        auto_btn.clicked.connect(self.auto_assign_dialog)
        btn_row.addWidget(refresh)
        btn_row.addWidget(edit_team)
        btn_row.addWidget(manage)
        btn_row.addWidget(delete_team)
        btn_row.addWidget(auto_btn)
        v.addLayout(btn_row)
        self.team_table.cellDoubleClicked.connect(lambda r, c: self.open_manage_dialog())
        self.load_teams()
        return w

    def create_team(self):
        name = self.team_name_in.text().strip()
        comp = self.team_comp_in.currentText()
        vet = 1 if self.team_vet_in.isChecked() else 0
        if not name:
            QMessageBox.warning(self, "Erro", "Nome da equipe é obrigatório")
            return
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("INSERT INTO teams (name, competition, is_veteran) VALUES (?,?,?)", (name, comp, vet))
        conn.commit()
        conn.close()
        self.team_name_in.clear()
        self.team_vet_in.setChecked(False)
        self.load_teams()
        # atualizar combos
        try:
            fill_team_combobox(self.eval_team_cb)
            fill_team_combobox(self.a_team_cb)
            fill_team_combobox(self.d_team_cb)
        except Exception:
            pass

    def load_teams(self):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT id,name,competition,is_veteran FROM teams ORDER BY id DESC")
        rows = cur.fetchall()
        conn.close()
        self.team_table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            tid, name, comp, vet = row
            self.team_table.setItem(r, 0, QTableWidgetItem(str(tid)))
            self.team_table.setItem(r, 1, QTableWidgetItem(str(name)))
            self.team_table.setItem(r, 2, QTableWidgetItem(str(comp)))
            self.team_table.setItem(r, 3, QTableWidgetItem("Sim" if vet else "Não"))
        try:
            self.team_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
            self.team_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
            self.team_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        except Exception:
            pass

    def delete_selected_team(self):
        sel = self.team_table.currentRow()
        if sel < 0:
            QMessageBox.warning(self, "Erro", "Selecione uma equipe para remover")
            return
        item = self.team_table.item(sel,0)
        if not item:
            return
        tid = int(item.text())
        ok = QMessageBox.question(self, "Confirmar", f"Remover equipe {tid}? Membros serão desvinculados.")
        if ok != QMessageBox.Yes:
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("DELETE FROM team_members WHERE team_id=?", (tid,))
        c.execute("DELETE FROM teams WHERE id=?", (tid,))
        conn.commit()
        conn.close()
        self.load_teams()

    def edit_selected_team(self):
        if get_process_status() == "ENCERRADO":
            QMessageBox.warning(self, "Ação Bloqueada", "Processo encerrado. Não é possível editar equipes.")
            return
        team_id = self._get_selected_id(self.team_table, "equipe")
        if team_id is None:
            return
        dlg = TeamEditDialog(team_id, parent=self)
        if dlg.exec() == QDialog.Accepted:
            self.load_teams()
            try:
                fill_team_combobox(self.eval_team_cb)
                fill_team_combobox(self.a_team_cb)
                fill_team_combobox(self.d_team_cb)
            except Exception:
                pass

    def open_manage_dialog(self):
        team_id = None
        sel = self.team_table.currentRow() if hasattr(self, 'team_table') else -1
        if sel >= 0:
            item = self.team_table.item(sel,0)
            if item:
                try:
                    team_id = int(item.text())
                except Exception:
                    team_id = None
        if not team_id:
            team_id_str, ok = QInputDialog.getText(self, "Gerenciar membros", "Insira Team ID:")
            if not ok:
                return
            try:
                team_id = int(team_id_str)
            except Exception:
                QMessageBox.warning(self, "Erro", "Team ID inválido")
                return
        dlg = TeamMemberDialog(team_id, parent=self)
        dlg.exec()
        self.load_teams()

    # --------------------------
    # PÁGINA: SESSÕES
    # --------------------------
    def page_sessions(self):
        w = QWidget()
        v = QVBoxLayout(w)
        form = QFormLayout()
        self.s_date = QLineEdit(datetime.now().strftime("%Y-%m-%d"))
        self.s_start = QLineEdit("07:00")
        self.s_end   = QLineEdit("12:20")
        form.addRow("Data (YYYY-MM-DD):", self.s_date)
        form.addRow("Início (HH:MM):", self.s_start)
        form.addRow("Fim (HH:MM):", self.s_end)
        btn = QPushButton("Criar sessão")
        btn.setObjectName("primary")
        btn.clicked.connect(self.create_session)
        v.addLayout(form)
        v.addWidget(btn)
        self.session_table = QTableWidget(0, 4)
        self.session_table.setHorizontalHeaderLabels(["ID", "Data", "Início", "Fim"])
        v.addWidget(self.session_table)
        btns = QHBoxLayout()
        refresh = QPushButton("Atualizar")
        refresh.setObjectName("primary")
        refresh.clicked.connect(self.load_sessions)
        edit_btn = QPushButton("Editar sessão")
        edit_btn.setObjectName("primary")
        edit_btn.clicked.connect(self.edit_selected_session)
        delete_btn = QPushButton("Remover sessão")
        delete_btn.setObjectName("danger")
        delete_btn.clicked.connect(self.delete_selected_session)
        btns.addWidget(refresh)
        btns.addWidget(edit_btn)
        btns.addWidget(delete_btn)
        v.addLayout(btns)
        self.load_sessions()
        return w

    def create_session(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("INSERT INTO training_sessions (date,start_time,end_time) VALUES (?,?,?)",
                  (self.s_date.text().strip(), self.s_start.text().strip(), self.s_end.text().strip()))
        conn.commit(); conn.close()
        QMessageBox.information(self, "OK", "Sessão criada")
        self.load_sessions()
        try:
            fill_session_combobox(self.eval_session_cb)
            fill_session_combobox(self.a_session_cb)
        except Exception:
            pass

    def load_sessions(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT id,date,start_time,end_time FROM training_sessions ORDER BY id DESC")
        rows = c.fetchall()
        conn.close()
        self.session_table.setRowCount(len(rows))
        for r,row in enumerate(rows):
            for cidx,val in enumerate(row):
                self.session_table.setItem(r,cidx,QTableWidgetItem(str(val)))

    def edit_selected_session(self):
        if get_process_status() == "ENCERRADO":
            QMessageBox.warning(self, "Ação Bloqueada", "Processo encerrado. Não é possível editar sessões.")
            return
        session_id = self._get_selected_id(self.session_table, "sessão")
        if session_id is None:
            return
        dlg = SessionEditDialog(session_id, parent=self)
        if dlg.exec() == QDialog.Accepted:
            self.load_sessions()
            try:
                fill_session_combobox(self.eval_session_cb)
                fill_session_combobox(self.a_session_cb)
            except Exception:
                pass

    def delete_selected_session(self):
        if get_process_status() == "ENCERRADO":
            QMessageBox.warning(self, "Ação Bloqueada", "Processo encerrado. Não é possível excluir sessões.")
            return
        session_id = self._get_selected_id(self.session_table, "sessão")
        if session_id is None:
            return
        ok = QMessageBox.question(
            self,
            "Confirmar",
            f"Remover sessão {session_id}? Presenças e avaliações associadas permanecerão.",
        )
        if ok != QMessageBox.Yes:
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("DELETE FROM training_sessions WHERE id=?", (session_id,))
        conn.commit()
        conn.close()
        audit('session_delete', f'session_id={session_id}')
        self.load_sessions()
        try:
            fill_session_combobox(self.eval_session_cb)
            fill_session_combobox(self.a_session_cb)
        except Exception:
            pass

    # --------------------------
    # PÁGINA: PRESENÇA (IDs via ComboBox)
    # --------------------------
    def page_attendance(self):
        w = QWidget()
        v = QVBoxLayout(w)
        form = QFormLayout()
        self.a_session_cb = QComboBox(); fill_session_combobox(self.a_session_cb)
        self.a_team_cb = QComboBox(); fill_team_combobox(self.a_team_cb)
        self.a_present = QComboBox(); self.a_present.addItems(["Sim", "Não"])
        self.a_notes = QLineEdit("")
        form.addRow("Sessão:", self.a_session_cb)
        form.addRow("Equipe:", self.a_team_cb)
        form.addRow("Presente?", self.a_present)
        form.addRow("Notas:", self.a_notes)
        btn = QPushButton("Registrar presença")
        btn.setObjectName("primary")
        btn.clicked.connect(self.add_attendance)
        v.addLayout(form)
        v.addWidget(btn)
        self.att_table = QTableWidget(0, 5)
        self.att_table.setHorizontalHeaderLabels(["ID","Sessão","Equipe","Presente","Notas"])
        v.addWidget(self.att_table)
        refresh = QPushButton("Atualizar")
        refresh.setObjectName("primary")
        refresh.clicked.connect(self.load_attendance)
        edit_btn = QPushButton("Editar presença")
        edit_btn.setObjectName("primary")
        edit_btn.clicked.connect(self.edit_selected_attendance)
        delete_btn = QPushButton("Remover presença")
        delete_btn.setObjectName("danger")
        delete_btn.clicked.connect(self.delete_selected_attendance)
        btns = QHBoxLayout()
        btns.addWidget(refresh)
        btns.addWidget(edit_btn)
        btns.addWidget(delete_btn)
        v.addLayout(btns)
        self.load_attendance()
        return w

    def add_attendance(self):
        sid = self.a_session_cb.currentData()
        tid = self.a_team_cb.currentData()
        if sid is None or tid is None:
            QMessageBox.warning(self, "Erro", "Selecione sessão e equipe")
            return
        pres = 1 if self.a_present.currentText() == "Sim" else 0
        notes = self.a_notes.text().strip()
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("INSERT INTO attendance (training_session_id,team_id,present,notes) VALUES (?,?,?,?)",
                  (sid, tid, pres, notes))
        conn.commit(); conn.close()
        QMessageBox.information(self, "OK", "Presença registrada")
        self.load_attendance()

    def load_attendance(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT id,training_session_id,team_id,present,notes FROM attendance ORDER BY id DESC")
        rows = c.fetchall()
        conn.close()
        self.att_table.setRowCount(len(rows))
        for r,row in enumerate(rows):
            rid, sid, tid, pres, notes = row
            self.att_table.setItem(r,0,QTableWidgetItem(str(rid)))
            self.att_table.setItem(r,1,QTableWidgetItem(str(sid)))
            self.att_table.setItem(r,2,QTableWidgetItem(str(tid)))
            self.att_table.setItem(r,3,QTableWidgetItem("Sim" if pres else "Não"))
            self.att_table.setItem(r,4,QTableWidgetItem(notes or ""))

    def edit_selected_attendance(self):
        if get_process_status() == "ENCERRADO":
            QMessageBox.warning(self, "Ação Bloqueada", "Processo encerrado. Não é possível editar presenças.")
            return
        attendance_id = self._get_selected_id(self.att_table, "presença")
        if attendance_id is None:
            return
        dlg = AttendanceEditDialog(attendance_id, parent=self)
        if dlg.exec() == QDialog.Accepted:
            self.load_attendance()

    def delete_selected_attendance(self):
        if get_process_status() == "ENCERRADO":
            QMessageBox.warning(self, "Ação Bloqueada", "Processo encerrado. Não é possível excluir presenças.")
            return
        attendance_id = self._get_selected_id(self.att_table, "presença")
        if attendance_id is None:
            return
        ok = QMessageBox.question(self, "Confirmar", f"Remover presença {attendance_id}?")
        if ok != QMessageBox.Yes:
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("DELETE FROM attendance WHERE id=?", (attendance_id,))
        conn.commit()
        conn.close()
        audit('attendance_delete', f'attendance_id={attendance_id}')
        self.load_attendance()

    # --------------------------
    # PÁGINA: AVALIAÇÕES (IDs via ComboBox)
    # --------------------------
    def page_evaluations(self):
        w = QWidget()
        v = QVBoxLayout(w)
        form = QFormLayout()
        self.eval_team_cb = QComboBox(); fill_team_combobox(self.eval_team_cb)
        self.eval_session_cb = QComboBox(); fill_session_combobox(self.eval_session_cb)
        self.eval_judge_in = QLineEdit()
        self.eval_imm_sb = QSpinBox(); self.eval_imm_sb.setRange(1,4); self.eval_imm_sb.setValue(3)
        self.eval_dev_sb = QSpinBox(); self.eval_dev_sb.setRange(1,4); self.eval_dev_sb.setValue(3)
        self.eval_pres_sb = QSpinBox(); self.eval_pres_sb.setRange(1,4); self.eval_pres_sb.setValue(3)
        self.eval_comment = QTextEdit()
        form.addRow("Equipe:", self.eval_team_cb)
        form.addRow("Sessão:", self.eval_session_cb)
        form.addRow("Banca (nome):", self.eval_judge_in)
        form.addRow("Imersão (1–4):", self.eval_imm_sb)
        form.addRow("Desenvolvimento (1–4):", self.eval_dev_sb)
        form.addRow("Apresentação (1–4):", self.eval_pres_sb)
        form.addRow("Comentário:", self.eval_comment)
        add_eval = QPushButton("Registrar avaliação")
        add_eval.setObjectName("primary")
        if get_process_status() == "ENCERRADO":
            add_eval.setDisabled(True)
            add_eval.setToolTip("Processo encerrado. Alterações não são mais permitidas.")
        add_eval.clicked.connect(self.add_evaluation)
        v.addLayout(form)
        v.addWidget(add_eval)
        self.recent_evals = QTableWidget(0, 7)
        self.recent_evals.setHorizontalHeaderLabels(["ID","Team","Sessão","Banca","Imersão","Desenv","Apres"])
        self.recent_evals.setSortingEnabled(True)
        self.recent_evals.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        v.addWidget(self.recent_evals)
        self.load_recent_evaluations()
        return w

    def add_evaluation(self):
        if get_process_status() == "ENCERRADO":
            QMessageBox.warning(self, "Ação Bloqueada", "Processo seletivo encerrado. Não é possível criar novas avaliações.")
            return
        team_id = self.eval_team_cb.currentData()
        session_id = self.eval_session_cb.currentData()
        if team_id is None or session_id is None:
            QMessageBox.warning(self, "Erro", "Selecione equipe e sessão")
            return
        judge = self.eval_judge_in.text().strip() or "Anon"
        imm = int(self.eval_imm_sb.value())
        dev = int(self.eval_dev_sb.value())
        pres = int(self.eval_pres_sb.value())
        comment = self.eval_comment.toPlainText().strip() or None
        
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        # Verificar avaliação duplicada para a mesma equipe na mesma sessão
        cur.execute("""
            SELECT id FROM evaluations
            WHERE team_id = ? AND training_session_id = ? AND is_active = 1
        """, (team_id, session_id))
        if cur.fetchone():
            QMessageBox.warning(self, "Erro de Validação", "Já existe uma avaliação ativa para esta equipe nesta sessão.")
            conn.close()
            return
        
        try:
            cur.execute("""
                INSERT INTO evaluations (team_id,judge,immersion,development,presentation,notes,training_session_id,comment)
                VALUES (?,?,?,?,?,?,?,?)
            """, (team_id, judge, imm, dev, pres, "", session_id, comment))
            evaluation_id = cur.lastrowid
            conn.commit()

            QMessageBox.information(self, "OK", f"Avaliação da equipe registrada (ID: {evaluation_id}).\nAgora, insira as contribuições individuais.")
            
            # Abrir diálogo de contribuição
            contrib_dlg = ContributionDialog(evaluation_id, team_id, self)
            contrib_dlg.exec()

        except Exception as e:
            QMessageBox.critical(self, "Erro de Banco de Dados", f"Não foi possível salvar a avaliação: {e}")
        finally:
            conn.close()

        self.eval_judge_in.clear(); self.eval_comment.clear()
        self.load_recent_evaluations()

    def load_recent_evaluations(self):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT id,team_id,training_session_id,judge,immersion,development,presentation FROM evaluations WHERE is_active = 1 ORDER BY id DESC LIMIT 50")
        rows = cur.fetchall()
        conn.close()
        self.recent_evals.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                self.recent_evals.setItem(r, c, QTableWidgetItem(str(val)))

    # --------------------------
    # PÁGINA: DIÁRIO (IDs via ComboBox)
    # --------------------------
    def page_diary(self):
        w = QWidget()
        v = QVBoxLayout(w)
        form = QFormLayout()
        self.d_team_cb = QComboBox(); fill_team_combobox(self.d_team_cb)
        self.d_title    = QLineEdit("Decisão técnica da semana")
        self.d_content  = QTextEdit()
        form.addRow("Equipe:", self.d_team_cb)
        form.addRow("Título:", self.d_title)
        form.addRow("Conteúdo:", self.d_content)
        btn_save = QPushButton("Registrar entrada")
        btn_save.setObjectName("primary")
        btn_save.clicked.connect(self.save_diary_entry)
        v.addLayout(form)
        v.addWidget(btn_save)
        attach_row = QHBoxLayout()
        self.attach_entry_id = QLineEdit("")  # preencher após criar entrada
        self.attach_btn_add  = QPushButton("Adicionar anexo à última entrada criada")
        self.attach_btn_add.setObjectName("primary")
        self.attach_btn_add.clicked.connect(self.add_attachment_last_entry)
        attach_row.addWidget(QLabel("Entry ID (auto):"))
        attach_row.addWidget(self.attach_entry_id)
        attach_row.addWidget(self.attach_btn_add)
        v.addLayout(attach_row)
        self.diary_table = QTableWidget(0, 4)
        self.diary_table.setHorizontalHeaderLabels(["ID","Equipe","Título","Criado em"])
        v.addWidget(self.diary_table)
        self.attach_table = QTableWidget(0, 4)
        self.attach_table.setHorizontalHeaderLabels(["ID","Entry ID","Arquivo","Nome original"])
        v.addWidget(self.attach_table)
        list_btns = QHBoxLayout()
        btn_load_d = QPushButton("Carregar entradas da equipe")
        btn_load_d.setObjectName("primary")
        btn_load_d.clicked.connect(self.load_diary_entries)
        btn_load_a = QPushButton("Carregar anexos da equipe")
        btn_load_a.setObjectName("primary")
        btn_load_a.clicked.connect(self.load_attachments_by_team)
        btn_edit_entry = QPushButton("Editar entrada selecionada")
        btn_edit_entry.setObjectName("primary")
        btn_edit_entry.clicked.connect(self.edit_selected_diary_entry)
        btn_delete_entry = QPushButton("Remover entrada selecionada")
        btn_delete_entry.setObjectName("danger")
        btn_delete_entry.clicked.connect(self.delete_selected_diary_entry)
        btn_delete_attach = QPushButton("Remover anexo selecionado")
        btn_delete_attach.setObjectName("danger")
        btn_delete_attach.clicked.connect(self.delete_selected_attachment)
        list_btns.addWidget(btn_load_d)
        list_btns.addWidget(btn_load_a)
        list_btns.addWidget(btn_edit_entry)
        list_btns.addWidget(btn_delete_entry)
        list_btns.addWidget(btn_delete_attach)
        v.addLayout(list_btns)
        return w

    def save_diary_entry(self):
        team_id = self.d_team_cb.currentData()
        if team_id is None:
            QMessageBox.warning(self, "Erro", "Selecione uma equipe")
            return
        title = self.d_title.text().strip()
        content = self.d_content.toPlainText().strip()
        if not title or not content:
            QMessageBox.warning(self, "Erro", "Título e conteúdo são obrigatórios")
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("INSERT INTO diary_entries (team_id,title,content,created_at) VALUES (?,?,?,?)",
                  (team_id, title, content, now_str()))
        conn.commit()
        c.execute("SELECT last_insert_rowid()")
        entry_id = c.fetchone()[0]
        conn.close()
        self.attach_entry_id.setText(str(entry_id))
        QMessageBox.information(self, "OK", f"Entrada criada (ID {entry_id})")
        self.d_title.clear(); self.d_content.clear()
        self.load_diary_entries()

    def add_attachment_last_entry(self):
        entry_id_text = self.attach_entry_id.text().strip()
        if not entry_id_text:
            QMessageBox.warning(self, "Erro", "Crie uma entrada primeiro")
            return
        entry_id = int(entry_id_text)
        fn, _ = QFileDialog.getOpenFileName(self, "Escolher arquivo")
        if not fn:
            return
        src = Path(fn)
        if not src.exists():
            QMessageBox.warning(self, "Erro", "Arquivo inexistente")
            return
        dst = ATTACH_DIR / f"{entry_id}_{src.name}"
        shutil.copy2(src, dst)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("INSERT INTO attachments (diary_entry_id,file_path,original_name,mime_type) VALUES (?,?,?,?)",
                  (entry_id, str(dst), src.name, ""))
        conn.commit(); conn.close()
        QMessageBox.information(self, "OK", "Anexo adicionado")
        self.load_attachments_by_team()

    def load_diary_entries(self):
        team_id = self.d_team_cb.currentData()
        if team_id is None:
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT id,team_id,title,created_at FROM diary_entries WHERE team_id=? ORDER BY id DESC", (team_id,))
        rows = c.fetchall()
        conn.close()
        self.diary_table.setRowCount(len(rows))
        for r,row in enumerate(rows):
            for cidx,val in enumerate(row):
                self.diary_table.setItem(r,cidx,QTableWidgetItem(str(val)))

    def load_attachments_by_team(self):
        team_id = self.d_team_cb.currentData()
        if team_id is None:
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("""
            SELECT a.id, a.diary_entry_id, a.file_path, a.original_name
            FROM attachments a
            JOIN diary_entries d ON d.id = a.diary_entry_id
            WHERE d.team_id=?
            ORDER BY a.id DESC
        """, (team_id,))
        rows = c.fetchall()
        conn.close()
        self.attach_table.setRowCount(len(rows))
        for r,row in enumerate(rows):
            for cidx,val in enumerate(row):
                self.attach_table.setItem(r,cidx,QTableWidgetItem(str(val)))

    def edit_selected_diary_entry(self):
        if get_process_status() == "ENCERRADO":
            QMessageBox.warning(self, "Ação Bloqueada", "Processo encerrado. Não é possível editar o diário.")
            return
        entry_id = self._get_selected_id(self.diary_table, "entrada")
        if entry_id is None:
            return
        dlg = DiaryEntryEditDialog(entry_id, parent=self)
        if dlg.exec() == QDialog.Accepted:
            self.load_diary_entries()
            self.load_attachments_by_team()

    def delete_selected_diary_entry(self):
        if get_process_status() == "ENCERRADO":
            QMessageBox.warning(self, "Ação Bloqueada", "Processo encerrado. Não é possível excluir o diário.")
            return
        entry_id = self._get_selected_id(self.diary_table, "entrada")
        if entry_id is None:
            return
        ok = QMessageBox.question(self, "Confirmar", f"Remover entrada {entry_id} e seus anexos?")
        if ok != QMessageBox.Yes:
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT file_path FROM attachments WHERE diary_entry_id=?", (entry_id,))
        attachments = [r[0] for r in c.fetchall()]
        c.execute("DELETE FROM attachments WHERE diary_entry_id=?", (entry_id,))
        c.execute("DELETE FROM diary_entries WHERE id=?", (entry_id,))
        conn.commit()
        conn.close()
        for path in attachments:
            try:
                Path(path).unlink(missing_ok=True)
            except Exception:
                pass
        audit('diary_entry_delete', f'entry_id={entry_id}')
        self.load_diary_entries()
        self.load_attachments_by_team()

    def delete_selected_attachment(self):
        if get_process_status() == "ENCERRADO":
            QMessageBox.warning(self, "Ação Bloqueada", "Processo encerrado. Não é possível excluir anexos.")
            return
        attachment_id = self._get_selected_id(self.attach_table, "anexo", col=0)
        if attachment_id is None:
            return
        ok = QMessageBox.question(self, "Confirmar", f"Remover anexo {attachment_id}?")
        if ok != QMessageBox.Yes:
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT id, file_path FROM attachments WHERE id=?", (attachment_id,))
        row = c.fetchone()
        if not row:
            conn.close()
            QMessageBox.warning(self, "Ação", "Anexo não encontrado.")
            return
        attach_id, file_path = row
        c.execute("DELETE FROM attachments WHERE id=?", (attach_id,))
        conn.commit()
        conn.close()
        try:
            Path(file_path).unlink(missing_ok=True)
        except Exception:
            pass
        audit('attachment_delete', f'attachment_id={attach_id}')
        self.load_attachments_by_team()

    # --------------------------
    # PÁGINA: DASHBOARD
    # --------------------------
    def page_dashboard(self):
        return Dashboard()

    # --------------------------
    # PÁGINA: SOBRE
    # --------------------------
    def page_about(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        
        about_text = QTextEdit()
        about_text.setReadOnly(True)
        
        content = """
        <h1>Sobre o Sistema de Processo Seletivo</h1>
        <p>Este aplicativo foi desenvolvido para gerenciar o processo seletivo da equipe de robótica <b>Robot One</b>, garantindo um processo justo, transparente e alinhado ao nosso modelo pedagógico.</p>

        <h2>1. Finalidade do Sistema</h2>
        <p>O objetivo principal é centralizar e organizar a avaliação de candidatos, com foco em três pilares:</p>
        <ul>
            <li><b>Trabalho em Equipe:</b> A unidade fundamental de avaliação é a equipe, não o indivíduo.</li>
            <li><b>Desenvolvimento Autônomo:</b> O sistema apoia a observação do progresso e da autonomia das equipes ao longo das sessões.</li>
            <li><b>Decisão Justa:</b> Fornecer dados consolidados e confiáveis para a banca avaliadora tomar suas decisões.</li>
        </ul>

        <h2>2. Modelo de Avaliação (Oficial)</h2>
        <p>A avaliação oficial, cujos resultados podem ser discutidos com as equipes, é sempre realizada no nível da <b>equipe</b>. Os critérios seguem o edital vigente:</p>
        <ul>
            <li><b>Imersão:</b> Engajamento e aprofundamento da equipe no problema proposto.</li>
            <li><b>Desenvolvimento:</b> A qualidade técnica e o progresso da solução criada.</li>
            <li><b>Apresentação:</b> Clareza e eficácia na comunicação das ideias e do projeto.</li>
        </ul>
        <p>A escala de notas (ex: 1 a 4) deve seguir o que foi publicamente divulgado no edital do processo seletivo.</p>

        <h2>3. Camada de Avaliação Interna</h2>
        <p>Para apoiar a decisão final da banca e identificar casos de "carona" ou de liderança excepcional, o sistema possui uma <b>camada de avaliação interna e oculta</b>. É fundamental compreender que:</p>
        <ul>
            <li>Esta camada de pesos individuais <b>não é exibida aos alunos</b> e não constitui uma nota individual pública.</li>
            <li>Seu propósito é exclusivamente para <b>análise e deliberação da banca</b>, permitindo uma decisão mais justa e detalhada.</li>
            <li>Os pesos são multiplicadores que ajustam o score da equipe para uma análise interna do desempenho de cada membro.</li>
        </ul>

        <h2>4. Boas Práticas de Uso</h2>
        <p>Para garantir a integridade dos dados, siga estas diretrizes:</p>
        <ul>
            <li><b>Registro Único:</b> Lance as notas de uma equipe apenas uma vez por sessão de avaliação para evitar duplicidade.</li>
            <li><b>Verificação:</b> Sempre confira se a equipe e a sessão selecionadas estão corretas antes de salvar uma avaliação.</li>
            <li><b>Diário de Bordo:</b> Utilize o Diário de Bordo para registrar observações técnicas, decisões de projeto e outras anotações relevantes sobre as equipes.</li>
            <li><b>Correções:</b> Qualquer necessidade de correção de notas ou desativação de uma avaliação deve ser feita <b>exclusivamente</b> pela área administrativa, com registro obrigatório do motivo.</li>
        </ul>

        <h2>5. Ética e Responsabilidade</h2>
        <p>Este sistema contém dados sensíveis e é uma ferramenta de grande responsabilidade.</p>
        <ul>
            <li><b>Confidencialidade:</b> Dados de avaliação interna e scores individuais são estritamente confidenciais e para uso exclusivo da banca.</li>
            <li><b>Não crie rankings públicos:</b> O sistema não deve ser usado para criar ou divulgar qualquer tipo de ranking de alunos. O objetivo é selecionar membros para uma equipe, não promover uma competição individual.</li>
            <li><b>Uso Correto:</b> O uso inadequado, a exposição de dados ou a manipulação de resultados comprometem a integridade e a justiça de todo o processo seletivo.</li>
        </ul>
        <br>
        <p><i>Lembre-se: o objetivo é formar a melhor equipe, e isso começa com um processo seletivo ético e bem conduzido.</i></p>
        """
        about_text.setHtml(content)
        
        layout.addWidget(about_text)
        return w


    # --------------------------
    # PÁGINA: ADMIN (oculto)
    # --------------------------
    def page_admin(self):
        w = QWidget()
        v = QVBoxLayout(w)
        v.addWidget(QLabel("Painel interno da banca — scores ocultos (sem ranking público)"))
        ops = QHBoxLayout()
        # view_btn = QPushButton("Ver avaliações")
        # view_btn.clicked.connect(self.show_admin_evaluations)
        calc_btn = QPushButton("Calcular scores ocultos")
        calc_btn.setObjectName("primary")
        calc_btn.clicked.connect(self.calculate_hidden_scores)
        dump_btn = QPushButton("Exportar avaliações (CSV)")
        dump_btn.setObjectName("primary")
        dump_btn.clicked.connect(self.export_evaluations)
        final_result_btn = QPushButton("Gerar Resultado Final")
        final_result_btn.setObjectName("primary")
        final_result_btn.clicked.connect(self.export_final_result)
        pin_btn = QPushButton("Trocar PIN")
        pin_btn.setObjectName("danger")
        pin_btn.clicked.connect(self.change_admin_pin)
        backup_btn = QPushButton("Backup DB")
        backup_btn.setObjectName("danger")
        backup_btn.clicked.connect(self.backup_db)
        # ops.addWidget(view_btn)
        ops.addWidget(calc_btn); ops.addWidget(dump_btn); ops.addWidget(final_result_btn)
        ops.addWidget(pin_btn); ops.addWidget(backup_btn)
        v.addLayout(ops)
        # Pesos internos
        wgt_box = QFormLayout()
        self.w_imm = QLineEdit("0.3"); self.w_dev = QLineEdit("0.5"); self.w_pres = QLineEdit("0.2")
        save_w = QPushButton("Salvar pesos internos")
        save_w.setObjectName("primary")
        if get_process_status() == "ENCERRADO":
            save_w.setDisabled(True)
            save_w.setToolTip("Processo encerrado. Alterações não são mais permitidas.")
        save_w.clicked.connect(self.save_internal_weights)
        wgt_box.addRow("Peso Imersão:", self.w_imm)
        wgt_box.addRow("Peso Desenvolvimento:", self.w_dev)
        wgt_box.addRow("Peso Apresentação:", self.w_pres)
        wgt_box.addRow(save_w)
        v.addLayout(wgt_box)

        # Controle de Estado do Processo Seletivo
        status_box = QFormLayout()
        self.process_status_cb = QComboBox()
        self.process_status_cb.addItems(["ABERTO", "ENCERRADO"])
        self.process_status_cb.setCurrentText(get_process_status())
        save_status_btn = QPushButton("Salvar Estado do Processo")
        save_status_btn.setObjectName("danger")
        save_status_btn.clicked.connect(self.change_process_status)
        status_box.addRow("Estado do Processo:", self.process_status_cb)
        status_box.addRow(save_status_btn)
        v.addLayout(status_box)

        # Tabela de avaliações
        self.admin_evals_table = QTableWidget(0, 11)
        self.admin_evals_table.setHorizontalHeaderLabels([
            "ID","Team","Sessão","Banca","Imersão","Desenv","Apres",
            "Hidden","Coment.","Ativo?","Ações"
        ])
        self.admin_evals_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.admin_evals_table.cellDoubleClicked.connect(self._admin_eval_cell_dbl)
        v.addWidget(self.admin_evals_table)
        # Resumo por equipe (ranking interno)
        self.chk_penalty = QCheckBox("Aplicar penalidade por presença (< 75% => x0.9)")
        self.chk_penalty.setChecked(True)
        btn_summary = QPushButton("Recalcular Resumo por equipe")
        btn_summary.setObjectName("primary")
        btn_summary.clicked.connect(self.recalc_team_summary)
        self.summary_table = QTableWidget(0, 5)
        self.summary_table.setHorizontalHeaderLabels(["Equipe ID", "Nome", "AVG hidden", "Presença (%)", "Score Final"])
        v.addWidget(self.chk_penalty)
        v.addWidget(btn_summary)
        v.addWidget(self.summary_table)
        
        # Resumo individual (ranking interno)
        v.addWidget(QLabel("Resumo de Contribuição Individual"))
        btn_ind_summary = QPushButton("Recalcular Resumo Individual")
        btn_ind_summary.setObjectName("primary")
        btn_ind_summary.clicked.connect(self.recalc_individual_summary)
        self.individual_summary_table = QTableWidget(0, 5)
        self.individual_summary_table.setHorizontalHeaderLabels(["Candidato ID", "Nome", "Equipe Atual", "Score Ponderado", "Avaliações"])
        v.addWidget(btn_ind_summary)
        v.addWidget(self.individual_summary_table)

        # Carrega pesos atuais e avaliações
        self.load_weights_into_form()
        self.load_admin_evaluations()
        return w

    def load_weights_into_form(self):
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        def get(name, default):
            c.execute("SELECT weight FROM internal_weights WHERE name=?", (name,))
            r = c.fetchone()
            return str(r[0]) if r else str(default)
        self.w_imm.setText(get('immersion', 0.3))
        self.w_dev.setText(get('development', 0.5))
        self.w_pres.setText(get('presentation', 0.2))
        conn.close()

    def save_internal_weights(self):
        try:
            wimm = float(self.w_imm.text())
            wdev = float(self.w_dev.text())
            wpres = float(self.w_pres.text())
        except Exception:
            QMessageBox.warning(self, "Erro", "Pesos devem ser números (float)")
            return
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        for name, w in [('immersion', wimm), ('development', wdev), ('presentation', wpres)]:
            c.execute("INSERT INTO internal_weights(name,weight) VALUES(?,?) ON CONFLICT(name) DO UPDATE SET weight=excluded.weight", (name, w))
        conn.commit(); conn.close()
        QMessageBox.information(self, "OK", "Pesos atualizados")
        audit('save_internal_weights', f"immersion={wimm},development={wdev},presentation={wpres}")

    def change_process_status(self):
        import hashlib
        cur = get_setting('admin_hash', '')
        pin, ok = QInputDialog.getText(self, 'Acesso Restrito', 'Insira o PIN da banca para alterar o estado do processo:', QLineEdit.Password)
        if not ok:
            return
        if hashlib.sha256(pin.encode()).hexdigest() != cur:
            QMessageBox.warning(self, 'Erro', 'PIN incorreto.')
            return

        selected_status = self.process_status_cb.currentText()
        set_process_status(selected_status)
        self.update_process_status_display()
        QMessageBox.information(self, "Sucesso", f"Estado do processo seletivo alterado para: {selected_status}")

    def load_admin_evaluations(self):
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("""
            SELECT id,team_id,training_session_id,judge,immersion,development,presentation,hidden_score,IFNULL(comment,''),
                   is_active, IFNULL(delete_reason,''), deleted_at
            FROM evaluations ORDER BY id DESC LIMIT 200
        """)
        rows = cur.fetchall()
        conn.close()
        self.admin_evals_table.setRowCount(len(rows))
        for r,row_data in enumerate(rows):
            is_active = row_data[9]
            delete_reason = row_data[10]
            is_permanently_deleted = delete_reason.startswith('[DELETED]')

            # Colore a linha
            row_color = None
            if is_permanently_deleted:
                row_color = Qt.darkRed
            elif not is_active:
                row_color = Qt.gray

            for c,val in enumerate(row_data[:9]): # Dados até a coluna 'Coment.'
                item = QTableWidgetItem(str(val))
                if row_color:
                    item.setBackground(row_color)
                self.admin_evals_table.setItem(r,c,item)

            # Coluna "Ativo?"
            active_item = QTableWidgetItem("Sim" if is_active else "Não")
            if row_color:
                active_item.setBackground(row_color)
            self.admin_evals_table.setItem(r, 9, active_item)

            # Coluna "Ações" com botões
            actions_widget = QWidget()
            actions_layout = QHBoxLayout(actions_widget)
            actions_layout.setContentsMargins(0,0,0,0)
            
            edit_btn = QPushButton("Editar")
            edit_btn.setObjectName("primary")
            edit_btn.clicked.connect(self._edit_evaluation_dialog)
            
            toggle_btn_text = "Desativar" if is_active else "Reativar"
            toggle_btn = QPushButton(toggle_btn_text)
            toggle_btn.setObjectName("primary")
            toggle_btn.clicked.connect(self._toggle_evaluation_active)

            delete_btn = QPushButton("Excluir")
            delete_btn.setObjectName("danger")
            delete_btn.setStyleSheet("background-color: #5d1b1b;")
            delete_btn.clicked.connect(self._delete_evaluation_logically)

            if get_process_status() == "ENCERRADO":
                edit_btn.setDisabled(True)
                edit_btn.setToolTip("Processo encerrado. Alterações não são mais permitidas.")
                toggle_btn.setDisabled(True)
                toggle_btn.setToolTip("Processo encerrado. Alterações não são mais permitidas.")
                delete_btn.setDisabled(True)
                delete_btn.setToolTip("Processo encerrado. Alterações não são mais permitidas.")

            actions_layout.addWidget(edit_btn)
            actions_layout.addWidget(toggle_btn)
            actions_layout.addWidget(delete_btn)

            if is_permanently_deleted:
                toggle_btn.setDisabled(True)
                delete_btn.setDisabled(True)
            
            self.admin_evals_table.setCellWidget(r, 10, actions_widget)
        self.admin_evals_table.resizeColumnsToContents()

    def _get_selected_eval_id(self):
        """Helper para pegar o ID da avaliação da linha selecionada na tabela admin."""
        sel_row = self.admin_evals_table.currentRow()
        if sel_row < 0:
            QMessageBox.warning(self, "Ação", "Selecione uma avaliação na tabela primeiro.")
            return None
        try:
            return int(self.admin_evals_table.item(sel_row, 0).text())
        except (ValueError, AttributeError):
            QMessageBox.critical(self, "Erro", "Não foi possível ler o ID da avaliação selecionada.")
            return None

    def _edit_evaluation_dialog(self):
        if get_process_status() == "ENCERRADO":
            QMessageBox.warning(self, "Ação Bloqueada", "Processo seletivo encerrado. Não é possível editar avaliações.")
            return
        eval_id = self._get_selected_eval_id()
        if eval_id is None:
            return
        
        dialog = EditEvaluationDialog(eval_id, self)
        if dialog.exec() == QDialog.Accepted:
            self.load_admin_evaluations()
            self.calculate_hidden_scores() # Recalcula após edição

    def _toggle_evaluation_active(self):
        if get_process_status() == "ENCERRADO":
            QMessageBox.warning(self, "Ação Bloqueada", "Processo seletivo encerrado. Não é possível ativar/desativar avaliações.")
            return
        eval_id = self._get_selected_eval_id()
        if eval_id is None:
            return

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT is_active, IFNULL(delete_reason, '') FROM evaluations WHERE id=?", (eval_id,))
        res = c.fetchone()
        conn.close()

        if not res:
            QMessageBox.critical(self, "Erro", f"Avaliação com ID {eval_id} não encontrada.")
            return

        is_currently_active, reason_text = res
        if reason_text.startswith('[DELETED]'):
            QMessageBox.warning(self, "Ação Bloqueada", "Esta avaliação foi permanentemente excluída e não pode ser reativada.")
            return

        new_status = 0 if is_currently_active else 1
        
        reason = ""
        if is_currently_active: # Desativando
            reason, ok = QInputDialog.getText(self, "Desativar Avaliação", "Motivo para desativar (obrigatório):")
            if not ok or not reason.strip():
                QMessageBox.warning(self, "Cancelado", "A desativação foi cancelada (motivo não fornecido).")
                return
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        if is_currently_active:
            c.execute(
                "UPDATE evaluations SET is_active=0, deleted_at=?, delete_reason=? WHERE id=?",
                (now_str(), reason, eval_id)
            )
            audit_action = 'evaluation_deactivate'
            audit_details = f"evaluation_id={eval_id}, reason='{reason}'"
        else: # Reativando
            c.execute("UPDATE evaluations SET is_active=1, deleted_at=NULL, delete_reason=NULL WHERE id=?", (eval_id,))
            audit_action = 'evaluation_reactivate'
            audit_details = f"evaluation_id={eval_id}"
        
        conn.commit()
        conn.close()
        
        audit(audit_action, audit_details)
        QMessageBox.information(self, "Sucesso", f"Avaliação {eval_id} foi {'desativada' if is_currently_active else 'reativada'}.")
        self.load_admin_evaluations()

    def _delete_evaluation_logically(self):
        if get_process_status() == "ENCERRADO":
            QMessageBox.warning(self, "Ação Bloqueada", "Processo seletivo encerrado. Não é possível excluir avaliações.")
            return
        eval_id = self._get_selected_eval_id()
        if eval_id is None:
            return

        ok = QMessageBox.warning(
            self, 
            "Confirmar Exclusão Lógica", 
            f"Você está prestes a excluir permanentemente a avaliação {eval_id}.\n"
            "Esta ação NÃO PODE ser desfeita pela interface.\n\n"
            "Deseja continuar?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if ok == QMessageBox.No:
            return

        reason, ok2 = QInputDialog.getText(self, "Justificativa Obrigatória", "Justifique a exclusão:")
        if not ok2 or not reason.strip():
            QMessageBox.warning(self, "Cancelado", "A exclusão foi cancelada (justificativa não fornecida).")
            return

        final_reason = f"[DELETED] {reason}"
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute(
            "UPDATE evaluations SET is_active=0, deleted_at=?, delete_reason=? WHERE id=?",
            (now_str(), final_reason, eval_id)
        )
        conn.commit()
        conn.close()

        audit('evaluation_logical_delete', f"evaluation_id={eval_id}, reason='{reason}'")
        QMessageBox.information(self, "Sucesso", f"Avaliação {eval_id} foi excluída logicamente.")
        self.load_admin_evaluations()


    def change_admin_pin(self):
        import hashlib
        cur = get_setting('admin_hash', '')
        cur_pin, ok = QInputDialog.getText(self, 'Trocar PIN', 'PIN atual:', QLineEdit.Password)
        if not ok:
            return
        if hashlib.sha256(cur_pin.encode()).hexdigest() != cur:
            QMessageBox.warning(self, 'Erro', 'PIN atual incorreto')
            return
        new_pin, ok2 = QInputDialog.getText(self, 'Trocar PIN', 'Novo PIN (4 dígitos):', QLineEdit.Password)
        if not ok2 or not new_pin:
            return
        confirm, ok3 = QInputDialog.getText(self, 'Trocar PIN', 'Confirme novo PIN:', QLineEdit.Password)
        if not ok3 or confirm != new_pin:
            QMessageBox.warning(self, 'Erro', 'Confirmação não confere')
            return
        h = hashlib.sha256(new_pin.encode()).hexdigest()
        set_setting('admin_hash', h)
        QMessageBox.information(self, 'OK', 'PIN alterado')
        audit('change_admin_pin', 'admin PIN changed')

    def backup_db(self):
        dst = Path(f"backup_selection_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
        shutil.copy2(DB_PATH, dst)
        QMessageBox.information(self, 'OK', f'Backup criado: {dst.name}')
        audit('backup_db', str(dst))

    def show_admin_evaluations(self):
        self.load_admin_evaluations()

    def calculate_hidden_scores(self):
        # Score oculto ponderado pelos pesos internos
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        weights = {}
        for name in ('immersion','development','presentation'):
            c.execute("SELECT weight FROM internal_weights WHERE name=?", (name,))
            r = c.fetchone(); weights[name] = float(r[0] if r else 1.0)
        c.execute("SELECT id,immersion,development,presentation FROM evaluations WHERE is_active = 1")
        rows = c.fetchall()
        for eid, imm, dev, pres in rows:
            imm = imm or 0; dev = dev or 0; pres = pres or 0
            hs = imm*weights['immersion'] + dev*weights['development'] + pres*weights['presentation']
            c.execute("UPDATE evaluations SET hidden_score=? WHERE id=?", (hs, eid))
        conn.commit(); conn.close()
        QMessageBox.information(self, "OK", "Scores ocultos recalculados para avaliações ativas.")
        self.load_admin_evaluations()
        audit('calculate_hidden_scores', 'recalculated using internal weights for active evaluations')

    def _admin_eval_cell_dbl(self, row, col):
        # permitir editar hidden_score no duplo clique
        if col == 7: # Hidden score column
            eval_id = self._get_selected_eval_id()
            if not eval_id: return

            item = self.admin_evals_table.item(row, 7)
            cur_val_str = item.text() if item else "0.0"
            try:
                cur_val = float(cur_val_str)
            except ValueError:
                cur_val = 0.0

            val, ok = QInputDialog.getDouble(self, "Editar Hidden Score Manualmente", "Novo score:", cur_val, 0.0, 9999.0, 3)
            if ok:
                reason, ok2 = QInputDialog.getText(self, "Justificativa", "Motivo da edição manual:")
                if ok2 and reason.strip():
                    conn = sqlite3.connect(DB_PATH)
                    c = conn.cursor()
                    c.execute("UPDATE evaluations SET hidden_score=? WHERE id=?", (val, eval_id))
                    conn.commit(); conn.close()
                    audit('manual_score_edit', f'eval_id={eval_id}, new_score={val}, reason="{reason}"')
                    self.load_admin_evaluations()
                else:
                    QMessageBox.warning(self, "Cancelado", "Edição cancelada (motivo não fornecido).")


    def export_evaluations(self):
        # Configuração do ranking
        APPROVED_COUNT = 5
        WAITLIST_COUNT = 5

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        try:
            # 1. Obter score base de cada avaliação (hidden_score)
            c.execute("SELECT id, hidden_score FROM evaluations WHERE is_active = 1")
            eval_scores = {eid: score for eid, score in c.fetchall()}
            if not eval_scores:
                QMessageBox.warning(self, "Exportar", "Nenhuma avaliação ativa encontrada para exportar.")
                return

            # 2. Obter contribuições individuais (pesos)
            c.execute("SELECT evaluation_id, member_id, weight FROM member_contribution")
            contributions = c.fetchall()

            # 3. Calcular score ponderado total por membro
            member_scores = {}
            for eval_id, member_id, weight in contributions:
                if member_id not in member_scores:
                    member_scores[member_id] = {'total_score': 0.0, 'eval_count': 0}
                
                team_score = eval_scores.get(eval_id, 0.0)
                member_scores[member_id]['total_score'] += team_score * (weight or 1.0)
                member_scores[member_id]['eval_count'] += 1
            
            if not member_scores:
                QMessageBox.warning(self, "Exportar", "Nenhuma contribuição individual encontrada para gerar o ranking.")
                return

            # 4. Obter dados dos candidatos (nome) e equipes
            c.execute("SELECT id, name FROM candidates")
            candidate_data = {cid: name for cid, name in c.fetchall()}
            
            # Subquery para pegar a equipe mais recente de um membro (ou uma qualquer)
            c.execute("""
                SELECT tm.candidate_id, t.name 
                FROM team_members tm 
                JOIN teams t ON tm.team_id = t.id
                GROUP BY tm.candidate_id
                HAVING tm.team_id = MAX(tm.team_id)
            """)
            member_teams = {cid: tname for cid, tname in c.fetchall()}

        finally:
            conn.close()

        # 5. Montar lista final para exportação
        summary_data = []
        for mid, data in member_scores.items():
            if data['eval_count'] > 0:
                summary_data.append({
                    'id': mid,
                    'name': candidate_data.get(mid, f'Candidato ID {mid}'),
                    'team': member_teams.get(mid, 'Sem equipe'),
                    'final_score': data['total_score']
                })
        
        # 6. Ordenar por score para gerar o ranking
        summary_data.sort(key=lambda x: x['final_score'], reverse=True)

        # 7. Escolher local para salvar
        default_filename = f"ranking_interno_{datetime.now().strftime('%Y%m%d')}.csv"
        fn, _ = QFileDialog.getSaveFileName(self, "Exportar Ranking Interno", default_filename, "CSV Files (*.csv)")

        if not fn:
            QMessageBox.information(self, "Exportar", "Exportação cancelada.")
            return

        # 8. Escrever o arquivo CSV
        try:
            import csv
            with open(fn, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Cabeçalho
                writer.writerow(['Posição no ranking', 'Nome do aluno', 'Equipe', 'Score final', 'Status final'])
                
                # Dados
                for idx, item in enumerate(summary_data):
                    rank = idx + 1
                    status = ''
                    if rank <= APPROVED_COUNT:
                        status = 'Aprovado'
                    elif rank <= APPROVED_COUNT + WAITLIST_COUNT:
                        status = 'Lista de espera'
                    else:
                        status = 'Não aprovado'
                    
                    writer.writerow([
                        rank,
                        item['name'],
                        item['team'],
                        f"{item['final_score']:.3f}",
                        status
                    ])
            
            QMessageBox.information(self, "Sucesso", f"Ranking interno exportado para\n{fn}")
            audit('export_ranking', f'file={Path(fn).name}, members={len(summary_data)}')
        except Exception as e:
            QMessageBox.critical(self, "Erro na Exportação", f"Não foi possível salvar o arquivo:\n{e}")

    def export_final_result(self):
        APPROVED_COUNT = 5 # Definido no README como exemplo
        WAITLIST_COUNT = 5 # Definido no README como exemplo

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        try:
            # 1. Obter score base de cada avaliação (hidden_score)
            c.execute("SELECT id, hidden_score FROM evaluations WHERE is_active = 1")
            eval_scores = {eid: score for eid, score in c.fetchall()}
            if not eval_scores:
                QMessageBox.warning(self, "Gerar Resultado Final", "Nenhuma avaliação ativa encontrada para calcular o resultado.")
                return

            # 2. Obter contribuições individuais (pesos)
            c.execute("SELECT evaluation_id, member_id, weight FROM member_contribution")
            contributions = c.fetchall()

            # 3. Calcular score ponderado total por membro
            member_scores = {}
            for eval_id, member_id, weight in contributions:
                if member_id not in member_scores:
                    member_scores[member_id] = {'total_score': 0.0, 'eval_count': 0}
                
                team_score = eval_scores.get(eval_id, 0.0)
                member_scores[member_id]['total_score'] += team_score * (weight or 1.0)
                member_scores[member_id]['eval_count'] += 1
            
            if not member_scores:
                QMessageBox.warning(self, "Gerar Resultado Final", "Nenhuma contribuição individual encontrada para gerar o resultado final.")
                return

            # 4. Obter dados dos candidatos (nome) e equipes
            c.execute("SELECT id, name FROM candidates")
            candidate_data = {cid: name for cid, name in c.fetchall()}
            
            # Subquery para pegar a equipe mais recente de um membro (ou uma qualquer)
            c.execute("""
                SELECT tm.candidate_id, t.name 
                FROM team_members tm 
                JOIN teams t ON tm.team_id = t.id
                GROUP BY tm.candidate_id
                HAVING tm.team_id = MAX(tm.team_id)
            """)
            member_teams = {cid: tname for cid, tname in c.fetchall()}

        finally:
            conn.close()

        # 5. Montar lista final para exportação
        summary_data = []
        for mid, data in member_scores.items():
            if data['eval_count'] > 0:
                summary_data.append({
                    'id': mid,
                    'name': candidate_data.get(mid, f'Candidato ID {mid}'),
                    'team': member_teams.get(mid, 'Sem equipe'),
                    'final_score': data['total_score']
                })
        
        # 6. Ordenar por score para gerar o ranking interno e definir status
        summary_data.sort(key=lambda x: x['final_score'], reverse=True)

        # 7. Escolher local para salvar
        default_filename = f"resultado_final_oficial_{datetime.now().strftime('%Y%m%d')}.csv"
        fn, _ = QFileDialog.getSaveFileName(self, "Gerar Resultado Final", default_filename, "CSV Files (*.csv)")

        if not fn:
            QMessageBox.information(self, "Gerar Resultado Final", "Geração de resultado final cancelada.")
            return

        # 8. Escrever o arquivo CSV
        try:
            import csv
            with open(fn, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Cabeçalho
                writer.writerow(['Nome do aluno', 'Equipe', 'Score final interno', 'Status final'])
                
                # Dados
                for idx, item in enumerate(summary_data):
                    rank = idx + 1
                    status = ''
                    if rank <= APPROVED_COUNT:
                        status = 'Aprovado'
                    elif rank <= APPROVED_COUNT + WAITLIST_COUNT:
                        status = 'Lista de espera'
                    else:
                        status = 'Não aprovado'
                    
                    writer.writerow([
                        item['name'],
                        item['team'],
                        f"{item['final_score']:.3f}",
                        status
                    ])
            
            QMessageBox.information(self, "Sucesso", f"Resultado final exportado para\n{fn}")
            audit('export_final_result', f'file={Path(fn).name}, members={len(summary_data)}')
        except Exception as e:
            QMessageBox.critical(self, "Erro na Geração", f"Não foi possível salvar o arquivo:\n{e}")


    # Resumo por equipe (ranking interno com penalidade opcional)
    def recalc_team_summary(self):
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute("""
            SELECT t.id, t.name, COALESCE(AVG(e.hidden_score), 0.0) AS avg_hidden
            FROM teams t
            LEFT JOIN evaluations e ON e.team_id = t.id AND e.is_active = 1
            GROUP BY t.id, t.name
            ORDER BY avg_hidden DESC
        """)
        rows = c.fetchall()
        c.execute("""
            SELECT team_id, AVG(present) AS pres_ratio
            FROM attendance
            GROUP BY team_id
        """)
        pres_map = {tid: (ratio or 0.0) for tid, ratio in c.fetchall()}
        conn.close()
        apply_penalty = self.chk_penalty.isChecked()
        out = []
        for tid, tname, avg_hidden in rows:
            pres_ratio = pres_map.get(tid, 0.0)
            final = avg_hidden
            if apply_penalty and pres_ratio < 0.75:
                final *= 0.9
            out.append((tid, tname, avg_hidden, pres_ratio*100.0, final))
        self.summary_table.setRowCount(len(out))
        for r, (tid, tname, avg_h, pres_pct, final) in enumerate(out):
            self.summary_table.setItem(r, 0, QTableWidgetItem(str(tid)))
            self.summary_table.setItem(r, 1, QTableWidgetItem(tname))
            self.summary_table.setItem(r, 2, QTableWidgetItem(f"{avg_h:.3f}"))
            self.summary_table.setItem(r, 3, QTableWidgetItem(f"{pres_pct:.1f}"))
            self.summary_table.setItem(r, 4, QTableWidgetItem(f"{final:.3f}"))

    def recalc_individual_summary(self):
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        
        # 1. Obter score de cada avaliação
        c.execute("SELECT id, hidden_score FROM evaluations WHERE is_active = 1")
        eval_scores = {eid: score for eid, score in c.fetchall()}

        # 2. Obter contribuições individuais
        c.execute("SELECT evaluation_id, member_id, weight FROM member_contribution")
        contributions = c.fetchall()

        # 3. Calcular score ponderado por membro
        member_scores = {}
        for eval_id, member_id, weight in contributions:
            if member_id not in member_scores:
                member_scores[member_id] = {'total_score': 0.0, 'eval_count': 0}
            
            team_score = eval_scores.get(eval_id, 0.0)
            member_scores[member_id]['total_score'] += team_score * (weight or 1.0)
            member_scores[member_id]['eval_count'] += 1

        # 4. Obter dados dos candidatos (nome, equipe atual)
        c.execute("SELECT id, name FROM candidates")
        candidate_data = {cid: name for cid, name in c.fetchall()}
        c.execute("SELECT tm.candidate_id, t.name FROM team_members tm JOIN teams t ON tm.team_id = t.id")
        # Pega a primeira equipe que encontrar para cada membro, para simplificar
        member_teams = {cid: tname for cid, tname in c.fetchall()}

        conn.close()

        # 5. Montar tabela de resultados
        summary_data = []
        for mid, data in member_scores.items():
            summary_data.append({
                'id': mid,
                'name': candidate_data.get(mid, 'N/A'),
                'team': member_teams.get(mid, 'Sem equipe'),
                'score': data['total_score'],
                'evals': data['eval_count']
            })
        
        # Ordenar por score
        summary_data.sort(key=lambda x: x['score'], reverse=True)

        self.individual_summary_table.setRowCount(len(summary_data))
        for r, item in enumerate(summary_data):
            self.individual_summary_table.setItem(r, 0, QTableWidgetItem(str(item['id'])))
            self.individual_summary_table.setItem(r, 1, QTableWidgetItem(item['name']))
            self.individual_summary_table.setItem(r, 2, QTableWidgetItem(item['team']))
            self.individual_summary_table.setItem(r, 3, QTableWidgetItem(f"{item['score']:.3f}"))
            self.individual_summary_table.setItem(r, 4, QTableWidgetItem(str(item['evals'])))
        
        audit('recalc_individual_summary', f'Calculated for {len(summary_data)} members')

    def save_contributions(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        try:
            # Check for existing contributions and update if found, otherwise insert
            for r in range(self.members_table.rowCount()):
                member_id = int(self.members_table.item(r, 0).text())
                weight = self.members_table.cellWidget(r, 2).value()
                note = self.members_table.cellWidget(r, 3).text().strip()
                
                c.execute("""
                    INSERT INTO member_contribution (evaluation_id, member_id, weight, note)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(evaluation_id, member_id) DO UPDATE SET weight=excluded.weight, note=excluded.note
                """, (self.evaluation_id, member_id, weight, note))
            conn.commit()
            QMessageBox.information(self, "Sucesso", "Contribuições individuais salvas/atualizadas.")
            audit('member_contribution_save', f'evaluation_id={self.evaluation_id}, count={self.members_table.rowCount()}')
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Erro de Banco de Dados", f"Não foi possível salvar as contribuições: {e}")
        finally:
            conn.close()

class EditEvaluationDialog(QDialog):
    def __init__(self, evaluation_id: int, parent=None):
        super().__init__(parent)
        self.evaluation_id = evaluation_id
        self.setWindowTitle(f"Editar Avaliação ID: {self.evaluation_id}")
        self.resize(400, 300)

        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.eval_imm_sb = QSpinBox(); self.eval_imm_sb.setRange(1,4)
        self.eval_dev_sb = QSpinBox(); self.eval_dev_sb.setRange(1,4)
        self.eval_pres_sb = QSpinBox(); self.eval_pres_sb.setRange(1,4)
        self.edit_reason_in = QLineEdit()

        form.addRow("Imersão (1–4):", self.eval_imm_sb)
        form.addRow("Desenvolvimento (1–4):", self.eval_dev_sb)
        form.addRow("Apresentação (1–4):", self.eval_pres_sb)
        form.addRow("Motivo da alteração:", self.edit_reason_in)

        layout.addLayout(form)

        save_btn = QPushButton("Salvar Alterações")
        save_btn.setObjectName("primary")
        save_btn.clicked.connect(self.save_changes)
        layout.addWidget(save_btn)

        self.load_evaluation_data()

    def load_evaluation_data(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT immersion, development, presentation FROM evaluations WHERE id=?", (self.evaluation_id,))
        data = c.fetchone()
        conn.close()
        if data:
            self.eval_imm_sb.setValue(data[0] or 0)
            self.eval_dev_sb.setValue(data[1] or 0)
            self.eval_pres_sb.setValue(data[2] or 0)

    def save_changes(self):
        reason = self.edit_reason_in.text().strip()
        if not reason:
            QMessageBox.warning(self, "Obrigatório", "O motivo da alteração é obrigatório.")
            return

        new_imm = self.eval_imm_sb.value()
        new_dev = self.eval_dev_sb.value()
        new_pres = self.eval_pres_sb.value()

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        # Log before changing
        c.execute("SELECT immersion, development, presentation FROM evaluations WHERE id=?", (self.evaluation_id,))
        old_data = c.fetchone()
        
        c.execute("""
            UPDATE evaluations
            SET immersion=?, development=?, presentation=?
            WHERE id=?
        """, (new_imm, new_dev, new_pres, self.evaluation_id))
        conn.commit()
        conn.close()

        details = (
            f"evaluation_id={self.evaluation_id}, reason='{reason}', "
            f"old_scores=({old_data[0]},{old_data[1]},{old_data[2]}), "
            f"new_scores=({new_imm},{new_dev},{new_pres})"
        )
        audit('evaluation_edit', details)

        QMessageBox.information(self, "Sucesso", "Avaliação atualizada.")
        self.accept()




class CandidateDialog(QDialog):
    def __init__(self, candidate_id:int, parent=None):
        super().__init__(parent)
        self.cid = candidate_id
        self.setWindowTitle(f"Candidato {candidate_id}")
        self.resize(640,550)

        layout = QVBoxLayout(self)

        # Formulário de edição
        form = QFormLayout()
        self.name_in = QLineEdit()
        self.email_in = QLineEdit()
        self.cpf_in = QLineEdit()
        self.phone_in = QLineEdit()
        self.grade_in = QLineEdit()
        self.notes_in = QTextEdit()
        form.addRow("Nome:", self.name_in)
        form.addRow("E-mail:", self.email_in)
        form.addRow("CPF:", self.cpf_in)
        form.addRow("Telefone:", self.phone_in)
        form.addRow("Série:", self.grade_in)
        form.addRow("Observações:", self.notes_in)
        layout.addLayout(form)

        save_btn = QPushButton("Salvar Alterações")
        save_btn.setObjectName("primary")
        save_btn.clicked.connect(self.save_data)
        layout.addWidget(save_btn)

        # Gerenciamento de equipes
        h = QHBoxLayout()
        col_left = QVBoxLayout()
        col_left.addWidget(QLabel("Pertence às equipes"))
        self.member_of = QListWidget()
        col_left.addWidget(self.member_of)
        rem = QPushButton("Remover da equipe selecionada")
        rem.setObjectName("danger")
        rem.clicked.connect(self.remove_from_team)   # <-- precisa deste método
        col_left.addWidget(rem)
        h.addLayout(col_left)

        col_right = QVBoxLayout()
        col_right.addWidget(QLabel("Equipes disponíveis"))
        self.available_teams = QListWidget()
        col_right.addWidget(self.available_teams)
        add = QPushButton("Adicionar à equipe selecionada")
        add.setObjectName("primary")
        add.clicked.connect(self.add_to_team)       # <-- e deste também
        col_right.addWidget(add)
        h.addLayout(col_right)

        layout.addLayout(h)

        self.load_data()

    def load_data(self):
        conn = sqlite3.connect(DB_PATH); c = conn.cursor()
        c.execute("SELECT name,email,notes,cpf,phone,grade FROM candidates WHERE id=?", (self.cid,))
        r = c.fetchone()
        if r:
            self.name_in.setText(r[0] or "")
            self.email_in.setText(r[1] or "")
            self.notes_in.setPlainText(r[2] or "")
            self.cpf_in.setText(r[3] or "")
            self.phone_in.setText(r[4] or "")
            self.grade_in.setText(r[5] or "")

        c.execute("SELECT t.id,t.name FROM teams t JOIN team_members m ON t.id=m.team_id WHERE m.candidate_id=?", (self.cid,))
        mems = c.fetchall()
        self.member_of.clear()
        for tid,tname in mems:
            self.member_of.addItem(f"{tid} - {tname}")

        c.execute("SELECT id,name FROM teams WHERE id NOT IN (SELECT team_id FROM team_members WHERE candidate_id=?)", (self.cid,))
        avail = c.fetchall()
        self.available_teams.clear()
        for tid,tname in avail:
            self.available_teams.addItem(f"{tid} - {tname}")
        conn.close()

    def save_data(self):
        name = self.name_in.text().strip()
        email = self.email_in.text().strip()
        notes = self.notes_in.toPlainText().strip()
        cpf = self.cpf_in.text().strip()
        phone = self.phone_in.text().strip()
        grade = self.grade_in.text().strip()
        if not name:
            QMessageBox.warning(self, "Erro", "Nome é obrigatório")
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("""
            UPDATE candidates SET
            name=?, email=?, notes=?, cpf=?, phone=?, grade=?
            WHERE id=?
        """, (name, email, notes, cpf, phone, grade, self.cid))
        conn.commit()
        conn.close()
        QMessageBox.information(self, "Sucesso", "Candidato atualizado.")
        self.accept()

    # >>> ADICIONADOS ABAIXO <<<
    def remove_from_team(self):
        selected_item = self.member_of.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "Erro", "Selecione uma equipe para remover.")
            return
        try:
            team_id = int(selected_item.text().split(" - ")[0])
        except Exception:
            QMessageBox.warning(self, "Erro", "Não foi possível ler o ID da equipe.")
            return

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        try:
            c.execute("DELETE FROM team_members WHERE team_id=? AND candidate_id=?", (team_id, self.cid))
            conn.commit()
            audit('team_member_remove', f'team={team_id}, candidate={self.cid}')
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Não foi possível remover: {e}")
        finally:
            conn.close()

        self.load_data()

    def add_to_team(self):
        selected_item = self.available_teams.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "Erro", "Selecione uma equipe para adicionar.")
            return
        try:
            team_id = int(selected_item.text().split(" - ")[0])
        except Exception:
            QMessageBox.warning(self, "Erro", "Não foi possível ler o ID da equipe.")
            return

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        try:
            c.execute("INSERT INTO team_members (team_id, candidate_id) VALUES (?, ?)", (team_id, self.cid))
            conn.commit()
            audit('team_member_add', f'team={team_id}, candidate={self.cid}')
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Erro", "Este candidato já está na equipe.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Não foi possível adicionar: {e}")
        finally:
            conn.close()

        self.load_data()

class TeamEditDialog(QDialog):
    def __init__(self, team_id: int, parent=None):
        super().__init__(parent)
        self.team_id = team_id
        self.setWindowTitle(f"Editar equipe {team_id}")
        self.resize(420, 220)

        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.name_in = QLineEdit()
        self.comp_in = QComboBox()
        self.comp_in.addItems(["OBR", "TBR", "CCBB"])
        self.vet_in = QCheckBox("Veteranos?")
        form.addRow("Nome:", self.name_in)
        form.addRow("Competição:", self.comp_in)
        form.addRow(self.vet_in)
        layout.addLayout(form)

        save_btn = QPushButton("Salvar")
        save_btn.setObjectName("primary")
        save_btn.clicked.connect(self.save_data)
        layout.addWidget(save_btn)
        self.load_data()

    def load_data(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT name, competition, is_veteran FROM teams WHERE id=?", (self.team_id,))
        row = c.fetchone()
        conn.close()
        if not row:
            QMessageBox.warning(self, "Erro", "Equipe não encontrada.")
            self.reject()
            return
        name, comp, vet = row
        self.name_in.setText(name or "")
        if comp:
            self.comp_in.setCurrentText(comp)
        self.vet_in.setChecked(bool(vet))

    def save_data(self):
        name = self.name_in.text().strip()
        comp = self.comp_in.currentText()
        vet = 1 if self.vet_in.isChecked() else 0
        if not name:
            QMessageBox.warning(self, "Erro", "Nome da equipe é obrigatório.")
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute(
            "UPDATE teams SET name=?, competition=?, is_veteran=? WHERE id=?",
            (name, comp, vet, self.team_id),
        )
        conn.commit()
        conn.close()
        audit('team_update', f'team_id={self.team_id}')
        QMessageBox.information(self, "Sucesso", "Equipe atualizada.")
        self.accept()

class SessionEditDialog(QDialog):
    def __init__(self, session_id: int, parent=None):
        super().__init__(parent)
        self.session_id = session_id
        self.setWindowTitle(f"Editar sessão {session_id}")
        self.resize(420, 200)

        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.date_in = QLineEdit()
        self.start_in = QLineEdit()
        self.end_in = QLineEdit()
        form.addRow("Data (YYYY-MM-DD):", self.date_in)
        form.addRow("Início (HH:MM):", self.start_in)
        form.addRow("Fim (HH:MM):", self.end_in)
        layout.addLayout(form)

        save_btn = QPushButton("Salvar")
        save_btn.setObjectName("primary")
        save_btn.clicked.connect(self.save_data)
        layout.addWidget(save_btn)
        self.load_data()

    def load_data(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT date, start_time, end_time FROM training_sessions WHERE id=?", (self.session_id,))
        row = c.fetchone()
        conn.close()
        if not row:
            QMessageBox.warning(self, "Erro", "Sessão não encontrada.")
            self.reject()
            return
        date, start, end = row
        self.date_in.setText(date or "")
        self.start_in.setText(start or "")
        self.end_in.setText(end or "")

    def save_data(self):
        date = self.date_in.text().strip()
        start = self.start_in.text().strip()
        end = self.end_in.text().strip()
        if not date or not start or not end:
            QMessageBox.warning(self, "Erro", "Data e horários são obrigatórios.")
            return
        try:
            datetime.strptime(date, "%Y-%m-%d")
            datetime.strptime(start, "%H:%M")
            datetime.strptime(end, "%H:%M")
        except ValueError:
            QMessageBox.warning(self, "Erro", "Formato inválido. Use YYYY-MM-DD e HH:MM.")
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute(
            "UPDATE training_sessions SET date=?, start_time=?, end_time=? WHERE id=?",
            (date, start, end, self.session_id),
        )
        conn.commit()
        conn.close()
        audit('session_update', f'session_id={self.session_id}')
        QMessageBox.information(self, "Sucesso", "Sessão atualizada.")
        self.accept()

class AttendanceEditDialog(QDialog):
    def __init__(self, attendance_id: int, parent=None):
        super().__init__(parent)
        self.attendance_id = attendance_id
        self.setWindowTitle(f"Editar presença {attendance_id}")
        self.resize(450, 220)

        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.session_cb = QComboBox()
        self.team_cb = QComboBox()
        for sid, d, s, e in fetch_sessions():
            self.session_cb.addItem(f"{sid} - {d} ({s}-{e})", sid)
        for tid, tname in fetch_teams():
            self.team_cb.addItem(f"{tid} - {tname}", tid)
        self.present_cb = QComboBox()
        self.present_cb.addItems(["Sim", "Não"])
        self.notes_in = QLineEdit()
        form.addRow("Sessão:", self.session_cb)
        form.addRow("Equipe:", self.team_cb)
        form.addRow("Presente?", self.present_cb)
        form.addRow("Notas:", self.notes_in)
        layout.addLayout(form)

        save_btn = QPushButton("Salvar")
        save_btn.setObjectName("primary")
        save_btn.clicked.connect(self.save_data)
        layout.addWidget(save_btn)
        self.load_data()

    def load_data(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute(
            "SELECT training_session_id, team_id, present, notes FROM attendance WHERE id=?",
            (self.attendance_id,),
        )
        row = c.fetchone()
        conn.close()
        if not row:
            QMessageBox.warning(self, "Erro", "Registro de presença não encontrado.")
            self.reject()
            return
        sid, tid, present, notes = row
        idx_session = self.session_cb.findData(sid)
        if idx_session >= 0:
            self.session_cb.setCurrentIndex(idx_session)
        idx_team = self.team_cb.findData(tid)
        if idx_team >= 0:
            self.team_cb.setCurrentIndex(idx_team)
        self.present_cb.setCurrentText("Sim" if present else "Não")
        self.notes_in.setText(notes or "")

    def save_data(self):
        sid = self.session_cb.currentData()
        tid = self.team_cb.currentData()
        present = 1 if self.present_cb.currentText() == "Sim" else 0
        notes = self.notes_in.text().strip()
        if sid is None or tid is None:
            QMessageBox.warning(self, "Erro", "Sessão e equipe são obrigatórias.")
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute(
            "UPDATE attendance SET training_session_id=?, team_id=?, present=?, notes=? WHERE id=?",
            (sid, tid, present, notes, self.attendance_id),
        )
        conn.commit()
        conn.close()
        audit('attendance_update', f'attendance_id={self.attendance_id}')
        QMessageBox.information(self, "Sucesso", "Presença atualizada.")
        self.accept()

class DiaryEntryEditDialog(QDialog):
    def __init__(self, entry_id: int, parent=None):
        super().__init__(parent)
        self.entry_id = entry_id
        self.setWindowTitle(f"Editar diário {entry_id}")
        self.resize(600, 420)

        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.team_cb = QComboBox()
        fill_team_combobox(self.team_cb)
        self.title_in = QLineEdit()
        self.content_in = QTextEdit()
        form.addRow("Equipe:", self.team_cb)
        form.addRow("Título:", self.title_in)
        form.addRow("Conteúdo:", self.content_in)
        layout.addLayout(form)

        save_btn = QPushButton("Salvar")
        save_btn.setObjectName("primary")
        save_btn.clicked.connect(self.save_data)
        layout.addWidget(save_btn)
        self.load_data()

    def load_data(self):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT team_id, title, content FROM diary_entries WHERE id=?", (self.entry_id,))
        row = c.fetchone()
        conn.close()
        if not row:
            QMessageBox.warning(self, "Erro", "Entrada não encontrada.")
            self.reject()
            return
        team_id, title, content = row
        idx = self.team_cb.findData(team_id)
        if idx >= 0:
            self.team_cb.setCurrentIndex(idx)
        self.title_in.setText(title or "")
        self.content_in.setPlainText(content or "")

    def save_data(self):
        team_id = self.team_cb.currentData()
        title = self.title_in.text().strip()
        content = self.content_in.toPlainText().strip()
        if team_id is None or not title or not content:
            QMessageBox.warning(self, "Erro", "Equipe, título e conteúdo são obrigatórios.")
            return
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute(
            "UPDATE diary_entries SET team_id=?, title=?, content=? WHERE id=?",
            (team_id, title, content, self.entry_id),
        )
        conn.commit()
        conn.close()
        audit('diary_entry_update', f'entry_id={self.entry_id}')
        QMessageBox.information(self, "Sucesso", "Entrada atualizada.")
        self.accept()

class ImportPreviewDialog(QDialog):
    def __init__(self, headers, preview_rows, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Pré-visualizar importação')
        self.resize(800, 400)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel('Revise as primeiras linhas e mapeie as colunas para Nome / E-mail / Observações'))
        cols = headers if headers else []
        col_count = len(cols) if cols else (max((len(r) for r in preview_rows), default=0))
        self.table = QTableWidget(len(preview_rows), col_count)
        if cols:
            self.table.setHorizontalHeaderLabels(cols)
        else:
            self.table.setHorizontalHeaderLabels([f'Col {i+1}' for i in range(col_count)])
        for r, row in enumerate(preview_rows):
            for c, val in enumerate(row):
                self.table.setItem(r, c, QTableWidgetItem(str(val) if val is not None else ''))
        layout.addWidget(self.table)
        form = QFormLayout()
        choices = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
        if not choices:
            choices = ['Col 1']
        self.map_name = QComboBox(); self.map_email = QComboBox(); self.map_notes = QComboBox()
        self.map_cpf = QComboBox(); self.map_phone = QComboBox(); self.map_grade = QComboBox()
        for ch in choices:
            self.map_name.addItem(ch)
            self.map_email.addItem(ch)
            self.map_notes.addItem(ch)
            self.map_cpf.addItem(ch)
            self.map_phone.addItem(ch)
            self.map_grade.addItem(ch)
        form.addRow('Coluna Nome:', self.map_name)
        form.addRow('Coluna E-mail:', self.map_email)
        form.addRow('Coluna CPF:', self.map_cpf)
        form.addRow('Coluna Telefone:', self.map_phone)
        form.addRow('Coluna Série:', self.map_grade)
        form.addRow('Coluna Observações:', self.map_notes)
        layout.addLayout(form)
        self.skip_dup = QCheckBox('Pular duplicados (mesmo nome+e-mail)')
        self.skip_dup.setChecked(True)
        layout.addWidget(self.skip_dup)
        btns = QHBoxLayout()
        ok = QPushButton('Confirmar')
        ok.setObjectName('primary')
        cancel = QPushButton('Cancelar')
        cancel.setObjectName('danger')
        ok.clicked.connect(self.accept)
        cancel.clicked.connect(self.reject)
        btns.addWidget(ok); btns.addWidget(cancel)
        layout.addLayout(btns)

    def mapping_indices(self):
        return (self.map_name.currentIndex(), self.map_email.currentIndex(), self.map_notes.currentIndex(), self.skip_dup.isChecked(),
                self.map_cpf.currentIndex(), self.map_phone.currentIndex(), self.map_grade.currentIndex())

class TeamMemberDialog(QDialog):
    def __init__(self, team_id: int, parent=None):
        super().__init__(parent)
        self.team_id = team_id
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT name FROM teams WHERE id=?", (self.team_id,))
        r = c.fetchone()
        team_name = r[0] if r else "Inválida"
        conn.close()

        self.setWindowTitle(f"Gerenciar Equipe {self.team_id} - {team_name}")
        self.resize(700, 500)
        layout = QHBoxLayout(self)

        # Coluna de membros atuais
        col_left = QVBoxLayout()
        col_left.addWidget(QLabel("Membros Atuais"))
        self.members_list = QListWidget()
        self.members_list.itemDoubleClicked.connect(self.remove_member)
        col_left.addWidget(self.members_list)
        self.remove_btn = QPushButton("<< Remover")
        self.remove_btn.setObjectName("danger")
        self.remove_btn.clicked.connect(self.remove_member)
        col_left.addWidget(self.remove_btn)
        layout.addLayout(col_left)

        # Coluna de candidatos disponíveis
        col_right = QVBoxLayout()
        col_right.addWidget(QLabel("Candidatos Disponíveis"))
        self.candidates_list = QListWidget()
        self.candidates_list.itemDoubleClicked.connect(self.add_member)
        col_right.addWidget(self.candidates_list)
        self.add_btn = QPushButton("Adicionar >>")
        self.add_btn.setObjectName("primary")
        self.add_btn.clicked.connect(self.add_member)
        col_right.addWidget(self.add_btn)
        layout.addLayout(col_right)

        self.load_data()

    def load_data(self):
        self.members_list.clear()
        self.candidates_list.clear()
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        # Carregar membros da equipe
        c.execute("""
            SELECT c.id, c.name FROM candidates c
            JOIN team_members tm ON c.id = tm.candidate_id
            WHERE tm.team_id = ? ORDER BY c.name
        """, (self.team_id,))
        for cid, name in c.fetchall():
            item = QListWidgetItem(f"{cid} - {name}")
            item.setData(Qt.UserRole, cid)
            self.members_list.addItem(item)

        # Carregar candidatos disponíveis
        c.execute("""
            SELECT id, name FROM candidates WHERE id NOT IN
            (SELECT candidate_id FROM team_members WHERE team_id = ?)
            ORDER BY name
        """, (self.team_id,))
        for cid, name in c.fetchall():
            item = QListWidgetItem(f"{cid} - {name}")
            item.setData(Qt.UserRole, cid)
            self.candidates_list.addItem(item)
        conn.close()

    
    def remove_member(self):
        selected_item = self.members_list.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "Erro", "Selecione um membro para remover.")
            return
        candidate_id = selected_item.data(Qt.UserRole)

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        try:
            c.execute("DELETE FROM team_members WHERE team_id=? AND candidate_id=?", (self.team_id, candidate_id))
            conn.commit()
            audit('team_member_remove', f'team={self.team_id}, candidate={candidate_id}')
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Não foi possível remover: {e}")
        finally:
            conn.close()
        self.load_data()

    def add_member(self):
        selected_item = self.candidates_list.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "Erro", "Selecione um candidato para adicionar.")
            return
        candidate_id = selected_item.data(Qt.UserRole)

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        try:
            c.execute("INSERT INTO team_members (team_id, candidate_id) VALUES (?, ?)", (self.team_id, candidate_id))
            conn.commit()
            audit('team_member_add', f'team={self.team_id}, candidate={candidate_id}')
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Erro", "Este candidato já está na equipe.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Não foi possível adicionar: {e}")
        finally:
            conn.close()
        self.load_data()


# --------------------------------
# MAIN
# --------------------------------
def main():
    init_db()
    # backup automático ao abrir
    backup_snapshot(prefix="startup")
    app = QApplication(sys.argv)
    with open("theme.qss", "r", encoding="utf-8") as f:
        app.setStyleSheet(f.read())
    f = app.font(); f.setPointSize(10); app.setFont(f)
    win = MainWindow()
    win.show()
    # backup ao fechar
    atexit.register(lambda: backup_snapshot(prefix="shutdown"))
    sys.exit(app.exec())

if __name__ == "__main__":
    main()

# Preciso adicionar uma dashboard mais limpa para maior visibilidade dos dados
